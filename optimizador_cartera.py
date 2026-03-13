import streamlit as st
import pandas as pd
import numpy as np
import yfinance as yf
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from scipy.optimize import minimize
import warnings
import json
import os
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────
#  PERSISTENCIA EN DISCO
# ─────────────────────────────────────────
DATA_FILE = "cartera_data.json"

def load_persistent():
    """Carga datos guardados en disco. Si no existe el archivo, devuelve defaults."""
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, 'r') as f:
                data = json.load(f)
            return data
        except:
            pass
    return None

def save_persistent(portfolio_df, instruments_df, sectors_list, sector_targets_dict=None):
    """Guarda cartera, instrumentos, sectores y targets de sector en disco."""
    try:
        data = {
            'portfolio': portfolio_df.to_dict(orient='records'),
            'instruments': instruments_df.to_dict(orient='records'),
            'sectors': sectors_list,
            'sector_targets': sector_targets_dict or {},
        }
        with open(DATA_FILE, 'w') as f:
            json.dump(data, f, ensure_ascii=False)
        return True
    except Exception as e:
        st.error(f"Error al guardar: {e}")
        return False

# ─────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────
st.set_page_config(
    page_title="Optimizador de Cartera",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
  .main { background-color: #0a0e1a; }
  .block-container { padding-top: 1.5rem; }
  h1, h2, h3 { color: #00d4ff; }
  .stTabs [data-baseweb="tab-list"] { gap: 6px; }
  .stTabs [data-baseweb="tab"] {
    background-color: #111827;
    border: 1px solid #1e2d45;
    border-radius: 8px;
    color: #64748b;
    padding: 6px 16px;
  }
  .stTabs [aria-selected="true"] {
    background-color: #00d4ff !important;
    color: #0a0e1a !important;
    font-weight: 700;
  }
  div[data-testid="metric-container"] {
    background: #111827;
    border: 1px solid #1e2d45;
    border-radius: 10px;
    padding: 12px 16px;
  }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────
#  SESSION STATE
# ─────────────────────────────────────────

# Cartera preseteada por defecto — el usuario puede modificarla y guardar cambios
DEFAULT_PORTFOLIO = pd.DataFrame([
    {'Ticker':'GOOGL','Monto_USD':25000,'Target_%':23,'Sector':'Comunicaciones'},
    {'Ticker':'ASML', 'Monto_USD':15000,'Target_%':14,'Sector':'Tecnología'},
    {'Ticker':'FXI',  'Monto_USD':12000,'Target_%':10,'Sector':'Brasil/EM'},
    {'Ticker':'EWZ',  'Monto_USD':10000,'Target_%':10,'Sector':'Brasil/EM'},
    {'Ticker':'XLV',  'Monto_USD':8000, 'Target_%':6, 'Sector':'Salud'},
    {'Ticker':'META', 'Monto_USD':6000, 'Target_%':8, 'Sector':'Comunicaciones'},
    {'Ticker':'MSFT', 'Monto_USD':4000, 'Target_%':5, 'Sector':'Tecnología'},
    {'Ticker':'IBM',  'Monto_USD':4000, 'Target_%':5.4,'Sector':'Tecnología'},
    {'Ticker':'PANW', 'Monto_USD':4000, 'Target_%':5, 'Sector':'Tecnología'},
    {'Ticker':'JPM',  'Monto_USD':3000, 'Target_%':2.8,'Sector':'Financiero'},
    {'Ticker':'V',    'Monto_USD':3000, 'Target_%':4, 'Sector':'Financiero'},
    {'Ticker':'RTX',  'Monto_USD':3000, 'Target_%':3, 'Sector':'Defensa'},
    {'Ticker':'CEG',  'Monto_USD':3000, 'Target_%':3.8,'Sector':'Energía'},
])

DEFAULT_SECTORS = [
    'Tecnología','Comunicaciones','Salud','Financiero',
    'Energía','Consumo','Industriales','Brasil/EM',
    'Minería/Cobre','Defensa','Otro'
]

# ── Inicializar session state cargando desde disco ──
if 'app_loaded' not in st.session_state:
    saved = load_persistent()
    if saved:
        st.session_state.portfolio       = pd.DataFrame(saved.get('portfolio', DEFAULT_PORTFOLIO.to_dict('records')))
        st.session_state.instruments     = pd.DataFrame(saved.get('instruments', DEFAULT_PORTFOLIO[['Ticker','Sector']].to_dict('records')))
        st.session_state.sectors         = saved.get('sectors', DEFAULT_SECTORS.copy())
        st.session_state.sector_targets  = saved.get('sector_targets', {})
    else:
        st.session_state.portfolio       = DEFAULT_PORTFOLIO.copy()
        st.session_state.instruments     = DEFAULT_PORTFOLIO[['Ticker','Sector']].copy().reset_index(drop=True)
        st.session_state.sectors         = DEFAULT_SECTORS.copy()
        st.session_state.sector_targets  = {}
    st.session_state.hist_data      = None
    st.session_state.tickers_loaded  = []
    st.session_state.app_loaded      = True
if 'balanz_data' not in st.session_state:
    st.session_state.balanz_data = None          # DataFrame importado de Balanz
if 'balanz_usd_tickers' not in st.session_state:
    st.session_state.balanz_usd_tickers = []     # Tickers con V.Actual en USD
if 'balanz_tipo_cambio' not in st.session_state:
    st.session_state.balanz_tipo_cambio = 1400.0 # Tipo de cambio default

def sync_instrument(ticker, sector=""):
    """Agrega un ticker al catálogo si no existe, y guarda en disco."""
    if ticker and ticker not in st.session_state.instruments['Ticker'].values:
        new_row = pd.DataFrame([{'Ticker': ticker, 'Sector': sector}])
        st.session_state.instruments = pd.concat(
            [st.session_state.instruments, new_row], ignore_index=True
        )
        save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))

# ─────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────
COLORS = px.colors.qualitative.Set2

def fmt_usd(v):
    return f"${v:,.0f}"

def fmt_pct(v):
    return f"{v:.2f}%"

@st.cache_data(ttl=3600, show_spinner=False)
def fetch_data(tickers, period="5y"):
    """Fetch historical closing prices from Yahoo Finance."""
    # Clean tickers — remove None, empty, whitespace-only
    tickers = [str(t).strip().upper() for t in tickers if t and str(t).strip()]
    if not tickers:
        st.error("No hay tickers válidos para descargar.")
        return None
    try:
        raw = yf.download(tickers, period=period, auto_adjust=True, progress=False)
        if isinstance(raw.columns, pd.MultiIndex):
            prices = raw['Close']
        else:
            prices = raw[['Close']] if 'Close' in raw.columns else raw
        # If only one ticker, ensure column name matches
        if len(tickers) == 1 and isinstance(prices, pd.DataFrame):
            prices.columns = tickers
        prices = prices.dropna(how='all')
        return prices
    except Exception as e:
        st.error(f"Error al descargar datos: {e}")
        return None

@st.cache_data(ttl=3600, show_spinner=False)
def validate_tickers(tickers):
    """Check which tickers exist on Yahoo Finance."""
    valid, invalid = [], []
    for t in tickers:
        try:
            info = yf.Ticker(t).fast_info
            if hasattr(info, 'last_price') and info.last_price and info.last_price > 0:
                valid.append(t)
            else:
                invalid.append(t)
        except:
            invalid.append(t)
    return valid, invalid

def calc_portfolio_weights(df):
    total = df['Monto_USD'].sum()
    df = df.copy()
    df['Peso_Actual_%'] = df['Monto_USD'] / total * 100 if total > 0 else 0
    df['Desvio_%'] = df['Peso_Actual_%'] - df['Target_%']
    df['Target_USD'] = df['Target_%'] / 100 * total
    df['Rebalanceo_USD'] = df['Target_USD'] - df['Monto_USD']
    df['Total'] = total
    return df

def calc_metrics(prices, weights=None, rf=0.02):
    """Calculate return, volatility, Sharpe, CAGR, Beta vs SPY."""
    prices = prices.dropna(how='all')
    if isinstance(prices, pd.Series):
        prices = prices.to_frame()
    if len(prices) < 30:
        raise ValueError("Datos insuficientes para calcular métricas (mínimo 30 días).")
    returns = prices.pct_change().dropna()
    if len(returns) == 0:
        raise ValueError("No hay retornos calculables con los datos disponibles.")
    annual_ret = returns.mean() * 252
    annual_vol = returns.std() * np.sqrt(252)
    sharpe = (annual_ret - rf) / annual_vol

    # CAGR — safe
    first = prices.iloc[0].replace(0, np.nan)
    last  = prices.iloc[-1]
    total_ret = last / first
    years = max(len(prices) / 252, 0.01)
    cagr = total_ret ** (1/years) - 1

    metrics = pd.DataFrame({
        'Retorno Anual %': (annual_ret * 100).round(2),
        'Volatilidad %': (annual_vol * 100).round(2),
        'Sharpe': sharpe.round(3),
        'CAGR %': (cagr * 100).round(2),
    })

    # Beta vs SPY
    try:
        spy_raw = yf.download('SPY', period=f'{max(int(years),1)}y', auto_adjust=True, progress=False)
        if isinstance(spy_raw.columns, pd.MultiIndex):
            spy = spy_raw['Close']['SPY'] if 'SPY' in spy_raw['Close'].columns else spy_raw['Close'].iloc[:,0]
        else:
            spy = spy_raw['Close'] if 'Close' in spy_raw.columns else spy_raw.iloc[:,0]
        spy = spy.squeeze()
        spy_ret = spy.pct_change().dropna()
        betas = {}
        for col in returns.columns:
            s1 = returns[col].dropna()
            s2 = spy_ret.dropna()
            common = s1.index.intersection(s2.index)
            if len(common) > 30:
                cov_matrix = np.cov(s1.loc[common].values, s2.loc[common].values)
                betas[col] = round(cov_matrix[0,1] / cov_matrix[1,1], 2) if cov_matrix[1,1] != 0 else np.nan
            else:
                betas[col] = np.nan
        metrics['Beta SPY'] = pd.Series(betas)
    except Exception as e:
        metrics['Beta SPY'] = np.nan

    # VaR 95%
    var_1d = (-returns.quantile(0.05) * 100).round(2)
    var_10d = (var_1d * np.sqrt(10)).round(2)
    metrics['VaR 1d 95%'] = var_1d
    metrics['VaR 10d 95%'] = var_10d

    return metrics, returns

def portfolio_metrics(weights_arr, returns_df, rf=0.02):
    port_ret = np.dot(weights_arr, returns_df.mean()) * 252
    port_vol = np.sqrt(np.dot(weights_arr, np.dot(returns_df.cov() * 252, weights_arr)))
    sharpe = (port_ret - rf) / port_vol
    return port_ret, port_vol, sharpe

def max_sharpe(returns_df, rf=0.02, min_weight=0.0):
    n = len(returns_df.columns)
    w0 = np.ones(n) / n
    lb = max(min_weight, 0.0)
    bounds = [(lb, 1.0)] * n
    constraints = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1}
    result = minimize(
        lambda w: -portfolio_metrics(w, returns_df, rf)[2],
        w0, method='SLSQP', bounds=bounds, constraints=constraints
    )
    return result.x if result.success else w0

def min_variance(returns_df, min_weight=0.0):
    n = len(returns_df.columns)
    w0 = np.ones(n) / n
    lb = max(min_weight, 0.0)
    bounds = [(lb, 1.0)] * n
    constraints = {'type': 'eq', 'fun': lambda w: np.sum(w) - 1}
    result = minimize(
        lambda w: portfolio_metrics(w, returns_df)[1],
        w0, method='SLSQP', bounds=bounds, constraints=constraints
    )
    return result.x if result.success else w0

def target_return_weights(returns_df, target_ret, min_weight=0.0):
    """Minimize volatility subject to hitting a target annual return."""
    n = len(returns_df.columns)
    w0 = np.ones(n) / n
    lb = max(min_weight, 0.0)
    bounds = [(lb, 1.0)] * n
    constraints = [
        {'type': 'eq', 'fun': lambda w: np.sum(w) - 1},
        {'type': 'eq', 'fun': lambda w: portfolio_metrics(w, returns_df)[0] - target_ret},
    ]
    result = minimize(
        lambda w: portfolio_metrics(w, returns_df)[1],
        w0, method='SLSQP', bounds=bounds, constraints=constraints
    )
    return result.x if result.success else None

def black_litterman(weights, returns_df, views_q, views_conf, rf=0.02, tau=0.05):
    """Simplified Black-Litterman model."""
    mu_eq = returns_df.mean() * 252
    sigma = returns_df.cov() * 252
    pi = mu_eq  # market equilibrium proxy

    n = len(weights)
    posterior_mu = pi.copy()
    for i, ticker in enumerate(returns_df.columns):
        if ticker in views_q:
            q = views_q[ticker]
            conf = views_conf.get(ticker, 0.5)
            posterior_mu[ticker] = (1 - conf) * pi[ticker] + conf * q

    return posterior_mu

def monte_carlo(port_ret, port_vol, total_usd, n_sims=500, n_days=252):
    dt = 1/252
    paths = np.zeros((n_sims, n_days+1))
    paths[:, 0] = total_usd
    z = np.random.standard_normal((n_sims, n_days))
    for d in range(1, n_days+1):
        paths[:, d] = paths[:, d-1] * np.exp(
            (port_ret - 0.5*port_vol**2)*dt + port_vol*np.sqrt(dt)*z[:, d-1]
        )
    return paths

# ─────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Configuración")

    # Años de historia — slider igual al otro optimizador
    years = st.slider("Años de historia", min_value=1, max_value=20, value=5,
        help="Cantidad de años de datos históricos a analizar")
    period_map = {1:"1y",2:"2y",3:"3y",4:"4y",5:"5y",6:"6y",7:"7y",8:"8y",
                  9:"9y",10:"10y",15:"15y",20:"20y"}
    PERIOD = period_map.get(years, f"{years}y")

    rf_rate = st.number_input("Tasa libre de riesgo (%)", value=2.0, step=0.25,
        help="Tasa anual libre de riesgo (ej: 2%)") / 100

    st.markdown("---")
    st.markdown("### 📐 Opciones avanzadas")

    # Retorno objetivo
    use_target_return = st.checkbox("Usar retorno objetivo",
        help="Optimizar la cartera buscando un retorno anual específico")
    target_return_pct = None
    if use_target_return:
        target_return_pct = st.number_input(
            "Retorno objetivo (%)", min_value=-50.0, max_value=100.0,
            value=10.0, step=0.5,
            help="Retorno anual objetivo para la optimización BL y Monte Carlo"
        )
        st.session_state['target_return'] = target_return_pct / 100
    else:
        st.session_state['target_return'] = None

    # Peso mínimo por activo
    min_weight_pct = st.number_input(
        "Peso mínimo por activo (%)", min_value=0.0, max_value=50.0,
        value=0.0, step=1.0,
        help="Peso mínimo que debe tener cada activo en la cartera optimizada (0 = sin mínimo)"
    )
    st.session_state['min_weight'] = min_weight_pct / 100

    st.markdown("---")
    st.markdown("### 🏷️ Gestión de sectores")
    new_sector = st.text_input("Nuevo sector", placeholder="Ej: REITs")
    if st.button("➕ Agregar sector") and new_sector.strip():
        if new_sector.strip() not in st.session_state.sectors:
            st.session_state.sectors.append(new_sector.strip())
            save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
            st.success(f"Sector '{new_sector}' agregado")
        else:
            st.warning("Ya existe ese sector")

    sector_to_del = st.selectbox("Eliminar sector", ["—"] + st.session_state.sectors)
    if st.button("🗑️ Eliminar sector") and sector_to_del != "—":
        st.session_state.sectors.remove(sector_to_del)
        save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
        st.success(f"Sector '{sector_to_del}' eliminado")

    st.markdown("---")
    if st.button("🚀 Cargar datos de mercado", type="primary", use_container_width=True):
        tickers = st.session_state.portfolio['Ticker'].tolist()
        if tickers:
            with st.spinner("Descargando datos de Yahoo Finance..."):
                st.session_state.hist_data = fetch_data(tickers, PERIOD)
                st.session_state.tickers_loaded = tickers
            st.success("✅ Datos cargados")
        else:
            st.warning("Primero cargá tu cartera")

    st.markdown("---")
    st.caption("📊 Optimizador de Cartera v4.0\nDatos: Yahoo Finance")

# ─────────────────────────────────────────
#  HEADER
# ─────────────────────────────────────────
st.markdown("# 📊 Optimizador de Cartera")
st.markdown("*Rebalanceo · Sectores · Correlación · Métricas · Black-Litterman · Monte Carlo — datos reales de Yahoo Finance*")
st.markdown("---")

# ─────────────────────────────────────────
#  TABS
# ─────────────────────────────────────────
tabs = st.tabs([
    "📊 Tablero Macro",
    "📥 Importar Balanz",
    "📋 Instrumentos",
    "📁 Mi Cartera",
    "🏷️ Sectores",
    "⚖️ Rebalanceo",
    "🎯 Estrategia",
    "🔗 Correlación",
    "📐 Métricas",
    "📈 Rendimiento",
    "⚠️ VaR",
    "🔥 Stress Test",
    "🌐 Frontera Eficiente",
    "🧠 Black-Litterman",
    "🎲 Monte Carlo",
])


# ══════════════════════════════════════════
#  TAB 0 — TABLERO MACRO
# ══════════════════════════════════════════
with tabs[0]:
    st.subheader("📊 Tablero Macro")

    if st.session_state.balanz_data is None:
        st.info("📥 Primero importá tu cartera desde la pestaña **📥 Importar Balanz**")
    else:
        df_b = st.session_state.balanz_data.copy()
        total_pesos = df_b['V_Actual_Pesos'].sum()

        if total_pesos == 0:
            st.error("Sin datos válidos")
        else:
            # ── KPIs principales ──
            st.markdown("### 💼 Cartera total")
            tc = st.session_state.balanz_tipo_cambio
            total_usd = total_pesos / tc
            k1,k2,k3 = st.columns(3)
            k1.metric("Total en Pesos", f"${total_pesos:,.0f}")
            k2.metric("Total en USD (aprox)", f"u$s {total_usd:,.0f}")
            k3.metric("Tipo de cambio usado", f"${tc:,.0f}")

            st.markdown("---")

            # ─────────────────────────────
            # BLOQUE 1 — MONEDA
            # ─────────────────────────────
            st.markdown("### 💱 Por Moneda")
            mon_group = df_b.groupby('Moneda_Display')['V_Actual_Pesos'].sum()
            mon_pct   = (mon_group / total_pesos * 100).round(1)

            col_m1, col_m2 = st.columns([1,1])
            with col_m1:
                fig_mon = go.Figure(go.Pie(
                    labels=mon_group.index, values=mon_group.values,
                    hole=0.45,
                    marker_colors=['#00d4ff','#f87171'],
                    textinfo='label+percent',
                    textfont=dict(color='#e2e8f0', size=13)
                ))
                fig_mon.update_layout(
                    paper_bgcolor='#111827', font_color='#e2e8f0',
                    title_font_color='#00d4ff', title_font_size=14,
                    title='Distribución por moneda',
                    showlegend=False, height=300, margin=dict(t=40,b=10)
                )
                st.plotly_chart(fig_mon, use_container_width=True)

            with col_m2:
                for mon, pct in mon_pct.items():
                    val = mon_group[mon]
                    st.metric(f"{'💵' if 'USD' in mon or 'Dólar' in mon else '🇦🇷'} {mon}",
                              f"{pct:.1f}%", f"${val:,.0f}")

            st.markdown("---")

            # ─────────────────────────────
            # BLOQUE 2 — TIPO DE RENTA con desglose moneda
            # ─────────────────────────────
            st.markdown("### 📈 Por Tipo de Renta")

            renta_types = ['Variable', 'Fija', 'Liquidez']
            renta_colors = {'Variable':'#00d4ff', 'Fija':'#4ade80', 'Liquidez':'#fbbf24'}

            # Normalize renta values
            def norm_renta(r):
                if r is None: return 'Otro'
                r = str(r).strip()
                if 'ariable' in r: return 'Variable'
                if 'ija' in r or 'ija' in r: return 'Fija'
                if 'iquidez' in r or 'iquid' in r or 'Plazo Fijo' in r: return 'Liquidez'
                return 'Otro'

            df_b['Renta_Norm'] = df_b['Renta'].apply(norm_renta)
            renta_group = df_b.groupby('Renta_Norm')['V_Actual_Pesos'].sum()
            renta_pct   = (renta_group / total_pesos * 100).round(1)

            col_r1, col_r2 = st.columns([1.2, 1])
            with col_r1:
                fig_renta = go.Figure(go.Bar(
                    x=list(renta_pct.index),
                    y=list(renta_pct.values),
                    marker_color=[renta_colors.get(r,'#94a3b8') for r in renta_pct.index],
                    text=[f"{v:.1f}%" for v in renta_pct.values],
                    textposition='outside', textfont=dict(color='#e2e8f0')
                ))
                fig_renta.update_layout(
                    paper_bgcolor='#111827', plot_bgcolor='#111827',
                    font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14,
                    title='% sobre cartera total', yaxis_title='%',
                    height=300, margin=dict(t=40,b=10), showlegend=False
                )
                st.plotly_chart(fig_renta, use_container_width=True)

            with col_r2:
                for renta in renta_group.index:
                    val     = renta_group[renta]
                    pct_tot = renta_pct[renta]
                    df_r    = df_b[df_b['Renta_Norm']==renta]

                    # Desglose por moneda dentro de este tipo de renta
                    mon_dentro = df_r.groupby('Moneda_Display')['V_Actual_Pesos'].sum()
                    total_r    = val

                    st.markdown(f"**{renta_colors.get(renta,'⬜')} {renta} — {pct_tot:.1f}% del total**")
                    for mon, mval in mon_dentro.items():
                        pct_dentro = mval / total_r * 100 if total_r > 0 else 0
                        pct_global = mval / total_pesos * 100
                        icon = '💵' if 'USD' in mon or 'Dólar' in mon else '🇦🇷'
                        st.markdown(f"&nbsp;&nbsp;&nbsp;{icon} {mon}: **{pct_dentro:.1f}%** de {renta} ({pct_global:.1f}% del total)")
                    st.markdown("---")

            # ─────────────────────────────
            # BLOQUE 3 — SECTOR MACRO (solo renta variable)
            # ─────────────────────────────
            st.markdown("### 🏭 Sectores MACRO (Renta Variable)")
            df_rv = df_b[df_b['Renta_Norm']=='Variable'].copy()
            if df_rv.empty:
                st.info("Sin instrumentos de renta variable")
            else:
                total_rv = df_rv['V_Actual_Pesos'].sum()
                sec_group = df_rv.groupby('Sector_Macro')['V_Actual_Pesos'].sum().sort_values(ascending=False)
                sec_pct_rv    = (sec_group / total_rv * 100).round(1)
                sec_pct_total = (sec_group / total_pesos * 100).round(1)

                col_s1, col_s2 = st.columns([1.2,1])
                with col_s1:
                    fig_sec = go.Figure(go.Bar(
                        y=list(sec_pct_rv.index),
                        x=list(sec_pct_rv.values),
                        orientation='h',
                        marker_color='rgba(0,212,255,0.7)',
                        text=[f"{v:.1f}%" for v in sec_pct_rv.values],
                        textposition='outside', textfont=dict(color='#e2e8f0')
                    ))
                    fig_sec.update_layout(
                        paper_bgcolor='#111827', plot_bgcolor='#111827',
                        font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14,
                        title='% sobre Renta Variable', height=350,
                        margin=dict(t=40,b=10), showlegend=False
                    )
                    st.plotly_chart(fig_sec, use_container_width=True)

                with col_s2:
                    for sec in sec_group.index:
                        pct_rv  = sec_pct_rv[sec]
                        pct_tot = sec_pct_total[sec]
                        st.markdown(f"**{sec}**: {pct_rv:.1f}% de RV ({pct_tot:.1f}% del total)")

            st.markdown("---")

            # ─────────────────────────────
            # BLOQUE 4 — TABLA DETALLE
            # ─────────────────────────────
            st.markdown("### 📋 Detalle por instrumento")
            df_show = df_b[['Ticker','Descripcion','Renta_Norm','Sector_Macro','Sector_Detalle',
                            'Moneda_Display','V_Actual_Pesos']].copy()
            df_show['% Cartera'] = (df_show['V_Actual_Pesos'] / total_pesos * 100).round(2)
            df_show = df_show.sort_values('V_Actual_Pesos', ascending=False)
            df_show.columns = ['Ticker','Descripción','Renta','Sector MACRO','Sector Detalle',
                               'Moneda','V. Actual ($)','% Cartera']
            st.dataframe(df_show, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════
#  TAB 1 — IMPORTAR BALANZ
# ══════════════════════════════════════════
with tabs[1]:
    st.subheader("📥 Importar desde Balanz")
    st.markdown("Copiá y pegá la info desde Balanz a tu Excel y subilo acá. La app lee la solapa **'Actualizar de Balanz'** automáticamente.")

    col_up, col_tc = st.columns([2,1])
    with col_up:
        uploaded_balanz = st.file_uploader(
            "Subí tu archivo Excel de Balanz (.xlsm o .xlsx)",
            type=['xlsm','xlsx','xls'],
            key="balanz_uploader"
        )
    with col_tc:
        tc_input = st.number_input(
            "Tipo de cambio ($ por u$s)",
            min_value=1.0, value=float(st.session_state.balanz_tipo_cambio),
            step=10.0, help="Se usa para convertir a pesos los instrumentos marcados como 'en dólares'"
        )
        st.session_state.balanz_tipo_cambio = tc_input

    if uploaded_balanz:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(uploaded_balanz, read_only=True, keep_vba=True, data_only=True)

            # Find the Balanz sheet
            sheet_name = None
            for name in wb.sheetnames:
                if 'balanz' in name.lower() or 'actualiz' in name.lower():
                    sheet_name = name
                    break
            if not sheet_name:
                st.error("No se encontró la solapa 'Actualizar de Balanz' en el archivo")
            else:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))

                # Find header row
                header_row = None
                for i, row in enumerate(rows[:10]):
                    if row and 'Ticker' in str(row):
                        header_row = i
                        break

                if header_row is None:
                    st.error("No se encontró el encabezado con 'Ticker' en la solapa")
                else:
                    headers = [str(h).strip() if h else '' for h in rows[header_row]]

                    # Map columns
                    col_map = {}
                    for i, h in enumerate(headers):
                        hl = h.lower()
                        if 'ticker' in hl: col_map['Ticker'] = i
                        elif 'v. actual' in hl or 'v.actual' in hl or 'actual' in hl and 'valor' not in hl: col_map['V_Actual'] = i
                        elif 'descripci' in hl: col_map['Descripcion'] = i
                        elif 'moneda' in hl: col_map['Moneda'] = i
                        elif 'sector macro' in hl or 'sector_macro' in hl or ('sector' in hl and 'macro' in hl): col_map['Sector_Macro'] = i
                        elif 'sector' in hl and 'macro' not in hl: col_map['Sector_Detalle'] = i
                        elif 'renta' in hl and 'rendimiento' not in hl: col_map['Renta'] = i
                        elif 'pa' in hl and 'is' in hl: col_map['Pais'] = i

                    # Also check for "V. Actual" specifically (col index 8 based on our reading)
                    if 'V_Actual' not in col_map:
                        for i, h in enumerate(headers):
                            if 'actual' in h.lower():
                                col_map['V_Actual'] = i
                                break

                    st.info(f"📋 Solapa encontrada: **{sheet_name}** | Columnas mapeadas: {list(col_map.keys())}")

                    # Parse data rows
                    data_rows = []
                    for row in rows[header_row+1:]:
                        if not row or not any(v is not None for v in row):
                            continue
                        ticker = str(row[col_map.get('Ticker', 0)] or '').strip()
                        if not ticker or ticker in ['#N/A','None','nan','']:
                            continue
                        # Get V.Actual — handle text like "u$s 117,85"
                        v_actual_raw = row[col_map.get('V_Actual', 8)]
                        try:
                            v_actual = float(str(v_actual_raw).replace('u$s','').replace('$','').replace(',','.').strip())
                        except:
                            continue
                        if v_actual <= 0:
                            continue

                        data_rows.append({
                            'Ticker':        ticker,
                            'Descripcion':   str(row[col_map.get('Descripcion',1)] or '').strip(),
                            'V_Actual_Raw':  v_actual,
                            'Moneda':        str(row[col_map.get('Moneda',22)] or '').strip(),
                            'Sector_Macro':  str(row[col_map.get('Sector_Macro',25)] or '').strip(),
                            'Sector_Detalle':str(row[col_map.get('Sector_Detalle',24)] or '').strip(),
                            'Renta':         str(row[col_map.get('Renta',18)] or '').strip(),
                            'Pais':          str(row[col_map.get('Pais',23)] or '').strip(),
                        })

                    if not data_rows:
                        st.error("No se encontraron datos válidos en la solapa")
                    else:
                        df_balanz = pd.DataFrame(data_rows)
                        st.success(f"✅ {len(df_balanz)} instrumentos leídos correctamente")

                        st.markdown("---")
                        st.markdown("### 💱 ¿Cuáles tienen el V. Actual en dólares?")
                        st.markdown("Tildá los instrumentos cuyo valor en la planilla está en **dólares** (no en pesos). Normalmente son fondos en USD o acciones compradas en el exterior.")

                        # Load previously saved USD tickers
                        prev_usd = st.session_state.balanz_usd_tickers

                        usd_selections = {}
                        cols_check = st.columns(4)
                        for i, ticker in enumerate(df_balanz['Ticker'].tolist()):
                            desc = df_balanz[df_balanz['Ticker']==ticker]['Descripcion'].values[0][:25]
                            val  = df_balanz[df_balanz['Ticker']==ticker]['V_Actual_Raw'].values[0]
                            with cols_check[i % 4]:
                                usd_selections[ticker] = st.checkbox(
                                    f"{ticker} ({val:,.0f})",
                                    value=(ticker in prev_usd),
                                    key=f"usd_check_{ticker}",
                                    help=desc
                                )

                        usd_tickers = [t for t, v in usd_selections.items() if v]

                        st.markdown("---")

                        if st.button("✅ Confirmar e importar cartera", type="primary"):
                            # Convert V.Actual to pesos
                            df_balanz['V_Actual_Pesos'] = df_balanz.apply(
                                lambda r: r['V_Actual_Raw'] * tc_input if r['Ticker'] in usd_tickers else r['V_Actual_Raw'],
                                axis=1
                            )
                            # Add display column for moneda
                            df_balanz['Moneda_Display'] = df_balanz['Ticker'].apply(
                                lambda t: 'Dólares (USD)' if t in usd_tickers else 'Pesos (ARS)'
                            )

                            st.session_state.balanz_data       = df_balanz
                            st.session_state.balanz_usd_tickers = usd_tickers

                            # Save USD tickers to persistent storage
                            save_persistent(
                                st.session_state.portfolio,
                                st.session_state.instruments,
                                st.session_state.sectors,
                                st.session_state.get('sector_targets', {})
                            )

                            total = df_balanz['V_Actual_Pesos'].sum()
                            st.success(f"✅ Cartera importada — {len(df_balanz)} instrumentos — Total: ${total:,.0f}")
                            st.info("👆 Ahora andá al **📊 Tablero Macro** para ver el análisis completo")
                            st.rerun()

        except Exception as e:
            st.error(f"Error al leer el archivo: {e}")
            import traceback
            st.code(traceback.format_exc())

    elif st.session_state.balanz_data is not None:
        df_b = st.session_state.balanz_data
        st.success(f"✅ Cartera importada — {len(df_b)} instrumentos — Total: ${df_b['V_Actual_Pesos'].sum():,.0f}")
        st.markdown(f"**Instrumentos con V.Actual en USD:** {', '.join(st.session_state.balanz_usd_tickers) or 'ninguno'}")
        if st.button("🗑️ Limpiar datos importados"):
            st.session_state.balanz_data = None
            st.rerun()


# ══════════════════════════════════════════
#  TAB 0 — INSTRUMENTOS (catálogo maestro)
# ══════════════════════════════════════════
with tabs[2]:
    st.subheader("📋 Catálogo de Instrumentos")
    st.markdown("Acá cargás tus instrumentos **una sola vez**. Quedan guardados y disponibles para armar tu cartera.")

    # ── Agregar instrumento al catálogo ──
    st.markdown("#### Agregar instrumento al catálogo")
    with st.form("add_instrument_form"):
        ci1, ci2, ci3 = st.columns([1.5, 2, 1])
        inst_ticker = ci1.text_input("Ticker", placeholder="AAPL").upper().strip()
        inst_sector = ci2.selectbox("Sector", [""] + st.session_state.sectors)
        inst_submit = ci3.form_submit_button("➕ Agregar", type="primary")

    if inst_submit and inst_ticker:
        import re
        if not re.match(r'^[A-Z0-9.\-]{1,10}$', inst_ticker):
            st.error("⚠️ Formato de ticker inválido")
        elif inst_ticker in st.session_state.instruments['Ticker'].values:
            st.warning(f"⚠️ **{inst_ticker}** ya está en el catálogo")
        else:
            with st.spinner(f"Verificando {inst_ticker} en Yahoo Finance..."):
                try:
                    info = yf.Ticker(inst_ticker).fast_info
                    ticker_ok = hasattr(info, 'last_price') and info.last_price and info.last_price > 0
                except:
                    ticker_ok = False
            if not ticker_ok:
                st.error(f"❌ **{inst_ticker}** no se encontró en Yahoo Finance")
            else:
                new_inst = pd.DataFrame([{'Ticker': inst_ticker, 'Sector': inst_sector}])
                st.session_state.instruments = pd.concat(
                    [st.session_state.instruments, new_inst], ignore_index=True
                )
                save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
                st.success(f"✅ **{inst_ticker}** agregado al catálogo")
                st.rerun()

    st.markdown("---")

    # ── Tabla del catálogo ──
    st.markdown(f"#### Mis instrumentos ({len(st.session_state.instruments)} cargados)")

    inst_edited = st.data_editor(
        st.session_state.instruments.copy(),
        column_config={
            'Ticker': st.column_config.TextColumn('Ticker', width='small', disabled=True),
            'Sector': st.column_config.SelectboxColumn('Sector', options=[""] + st.session_state.sectors, width='medium'),
        },
        hide_index=True,
        num_rows="dynamic",
        use_container_width=True,
        key="instruments_editor"
    )

    col_save, col_restore = st.columns([1,1])
    with col_save:
        if st.button("💾 Guardar cambios del catálogo", type="primary"):
            inst_edited_clean = inst_edited.dropna(subset=['Ticker'])
            inst_edited_clean['Ticker'] = inst_edited_clean['Ticker'].astype(str).str.upper().str.strip()
            inst_edited_clean = inst_edited_clean[inst_edited_clean['Ticker'].str.match(r'^[A-Z0-9.\-]{1,10}$', na=False)]
            st.session_state.instruments = inst_edited_clean.reset_index(drop=True)
            save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
            st.success("✅ Catálogo actualizado")
            st.rerun()
    with col_restore:
        if st.button("🔄 Restaurar catálogo original"):
            st.session_state.instruments = DEFAULT_PORTFOLIO[['Ticker','Sector']].copy().reset_index(drop=True)
            save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
            st.success("✅ Catálogo restaurado")
            st.rerun()

    st.info("💡 Los instrumentos del catálogo están disponibles en **Mi Cartera** para seleccionarlos con un click.")

# ══════════════════════════════════════════
#  TAB 1 — MI CARTERA
# ══════════════════════════════════════════
with tabs[3]:
    st.subheader("📁 Mi Cartera")

    # ── Agregar desde catálogo ──
    col1, col2 = st.columns([1.2, 1])
    with col1:
        st.markdown("#### Agregar desde catálogo de instrumentos")
        # Instruments not yet in portfolio
        in_portfolio = set(st.session_state.portfolio['Ticker'].values)
        available_inst = [
            t for t in st.session_state.instruments['Ticker'].values
            if t not in in_portfolio
        ]
        if available_inst:
            with st.form("add_from_catalog"):
                fc1, fc2, fc3 = st.columns([1.5, 1.2, 1])
                sel_ticker = fc1.selectbox("Instrumento", available_inst)
                sel_monto  = fc2.number_input("Monto USD", min_value=0.0, step=100.0, value=1000.0)
                sel_target = fc3.number_input("Target %", min_value=0.0, max_value=100.0, step=0.5)
                cat_submit = st.form_submit_button("➕ Agregar a mi cartera", type="primary")
            if cat_submit and sel_ticker:
                sel_sector = st.session_state.instruments[
                    st.session_state.instruments['Ticker']==sel_ticker
                ]['Sector'].values[0]
                new_row = pd.DataFrame([{
                    'Ticker': sel_ticker, 'Monto_USD': sel_monto,
                    'Target_%': sel_target, 'Sector': sel_sector
                }])
                st.session_state.portfolio = pd.concat(
                    [st.session_state.portfolio, new_row], ignore_index=True
                )
                save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
                st.success(f"✅ {sel_ticker} agregado a tu cartera")
                st.rerun()
        else:
            st.info("Todos los instrumentos del catálogo ya están en tu cartera.")

    with col2:
        st.markdown("#### Ticker nuevo (no está en catálogo)")
        with st.form("add_new_ticker"):
            n1, n2, n3 = st.columns([1.2, 1, 1])
            new_ticker = n1.text_input("Ticker", placeholder="AAPL").upper().strip()
            new_monto  = n2.number_input("Monto USD", min_value=0.0, step=100.0, value=1000.0)
            new_target = n3.number_input("Target %", min_value=0.0, max_value=100.0, step=0.5)
            new_sector = st.selectbox("Sector (se guardará en catálogo)", [""] + st.session_state.sectors)
            new_submit = st.form_submit_button("➕ Agregar y guardar en catálogo", type="primary")

        if new_submit and new_ticker:
            import re
            if not re.match(r'^[A-Z0-9.\-]{1,10}$', new_ticker):
                st.error("⚠️ Formato de ticker inválido")
            elif new_ticker in st.session_state.portfolio['Ticker'].values:
                st.warning(f"⚠️ {new_ticker} ya está en tu cartera")
            else:
                with st.spinner(f"Verificando {new_ticker} en Yahoo Finance..."):
                    try:
                        info = yf.Ticker(new_ticker).fast_info
                        ticker_ok = hasattr(info, 'last_price') and info.last_price and info.last_price > 0
                    except:
                        ticker_ok = False
                if not ticker_ok:
                    st.error(f"❌ **{new_ticker}** no se encontró en Yahoo Finance")
                else:
                    # Add to portfolio
                    new_row = pd.DataFrame([{
                        'Ticker': new_ticker, 'Monto_USD': new_monto,
                        'Target_%': new_target, 'Sector': new_sector
                    }])
                    st.session_state.portfolio = pd.concat(
                        [st.session_state.portfolio, new_row], ignore_index=True
                    )
                    # Sync to instruments catalog
                    sync_instrument(new_ticker, new_sector)
                    save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
                    st.success(f"✅ {new_ticker} agregado a tu cartera y al catálogo")
                    st.rerun()

    st.markdown("---")

    if st.button("🔄 Restaurar cartera de ejemplo"):
        st.session_state.portfolio = DEFAULT_PORTFOLIO.copy()
        save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
        st.rerun()

    if not st.session_state.portfolio.empty:
        df = calc_portfolio_weights(st.session_state.portfolio.copy())
        total = df['Total'].iloc[0]
        total_target = df['Target_%'].sum()

        # Metrics row
        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("Total cartera", fmt_usd(total))
        m2.metric("Activos", len(df))
        m3.metric("Total Target %", f"{total_target:.2f}%",
                  delta=f"{total_target-100:.2f}% vs 100%" if abs(total_target-100)>0.1 else "✓ OK")
        m4.metric("A comprar", len(df[df['Desvio_%']<-0.5]))
        m5.metric("A vender", len(df[df['Desvio_%']>0.5]))

        if abs(total_target - 100) > 0.1:
            st.warning(f"⚠️ Los targets suman {total_target:.2f}% — deben sumar 100%")

        # Editable table
        st.markdown("#### Editá tu cartera")

        # Sort controls
        sort_col_map = {
            'Monto USD': 'Monto_USD',
            'Peso Actual %': 'Peso_Actual_%',
            'Target %': 'Target_%',
            'Desvío %': 'Desvio_%',
            'Ticker': 'Ticker',
            'Sector': 'Sector',
        }
        sc1, sc2 = st.columns([3,1])
        sort_choice = sc1.selectbox("Ordenar por", list(sort_col_map.keys()), index=0, key="sort_portfolio")
        sort_asc = sc2.selectbox("Orden", ["↓ Mayor→Menor", "↑ Menor→Mayor"], key="sort_dir") == "↑ Menor→Mayor"
        sort_field = sort_col_map[sort_choice]
        df_sorted = df[['Ticker','Monto_USD','Target_%','Sector','Peso_Actual_%','Desvio_%']].round(2).sort_values(sort_field, ascending=sort_asc)

        edited = st.data_editor(
            df_sorted,
            column_config={
                'Ticker': st.column_config.TextColumn('Ticker', width='small'),
                'Monto_USD': st.column_config.NumberColumn('Monto USD', format="$%.0f"),
                'Target_%': st.column_config.NumberColumn('Target %', format="%.1f%%"),
                'Sector': st.column_config.SelectboxColumn('Sector', options=[""] + st.session_state.sectors),
                'Peso_Actual_%': st.column_config.NumberColumn('Peso Actual %', format="%.2f%%", disabled=True),
                'Desvio_%': st.column_config.NumberColumn('Desvío %', format="%.2f%%", disabled=True),
            },
            hide_index=True,
            num_rows="dynamic",
            use_container_width=True,
            key="portfolio_editor"
        )
        st.warning("⚠️ **Acordate de guardar** — cualquier cambio en la tabla recién se aplica al hacer click en el botón de abajo.")
        if st.button("💾 Guardar cambios en mi cartera", type="primary"):
            edited_clean = edited[['Ticker','Monto_USD','Target_%','Sector']].copy()
            edited_clean['Ticker'] = edited_clean['Ticker'].astype(str).str.upper().str.strip()
            edited_clean = edited_clean[edited_clean['Ticker'].str.match(r'^[A-Z0-9.\-]{1,10}$', na=False)]
            # Validate any NEW tickers not already in portfolio
            old_tickers = set(st.session_state.portfolio['Ticker'].values)
            new_tickers = [t for t in edited_clean['Ticker'].values if t not in old_tickers]
            invalid = []
            if new_tickers:
                with st.spinner(f"Validando tickers nuevos: {', '.join(new_tickers)}..."):
                    for t in new_tickers:
                        try:
                            info = yf.Ticker(t).fast_info
                            if not (hasattr(info, 'last_price') and info.last_price and info.last_price > 0):
                                invalid.append(t)
                        except:
                            invalid.append(t)
            if invalid:
                st.error(f"❌ Tickers no encontrados en Yahoo Finance: **{', '.join(invalid)}** — corregí o eliminá esas filas antes de guardar.")
            else:
                # Sync any new tickers to instruments catalog
                for _, row in edited_clean.iterrows():
                    sync_instrument(row['Ticker'], row.get('Sector',''))
                st.session_state.portfolio = edited_clean
                save_persistent(st.session_state.portfolio, st.session_state.instruments, st.session_state.sectors, st.session_state.get("sector_targets",{}))
                st.success("✅ Cambios guardados — nuevos tickers sincronizados al catálogo")
                st.rerun()

        col_exp1, col_exp2 = st.columns(2)
        with col_exp1:
            csv = df.to_csv(index=False).encode('utf-8')
            st.download_button("⬇️ Exportar CSV", csv, "mi_cartera.csv", "text/csv")
        with col_exp2:
            import io
            buf = io.BytesIO()
            df.to_excel(buf, index=False)
            st.download_button("⬇️ Exportar Excel", buf.getvalue(), "mi_cartera.xlsx")
    else:
        st.info("Cargá tu cartera usando el formulario o subiendo un Excel")


# ══════════════════════════════════════════
#  TAB 2 — SECTORES
# ══════════════════════════════════════════
with tabs[4]:
    st.subheader("🏷️ Análisis por Sector")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    else:
        df = calc_portfolio_weights(st.session_state.portfolio.copy())
        total = df['Total'].iloc[0]

        # Sector target editor
        st.markdown("#### Target % por sector (opcional)")
        if 'sector_targets' not in st.session_state:
            st.session_state.sector_targets = {s:0.0 for s in st.session_state.sectors}

        sector_cols = st.columns(min(4, len(st.session_state.sectors)))
        for i, s in enumerate(st.session_state.sectors):
            with sector_cols[i % len(sector_cols)]:
                val = st.number_input(s, min_value=0.0, max_value=100.0, step=1.0,
                    value=float(st.session_state.sector_targets.get(s,0)),
                    key=f"st_{s}")
                st.session_state.sector_targets[s] = val

        st.markdown("---")

        # Group by sector
        sector_group = df.groupby('Sector').agg(
            Monto_USD=('Monto_USD','sum'),
            Peso_Actual=('Peso_Actual_%','sum'),
            Tickers=('Ticker', lambda x: ', '.join(x))
        ).reset_index().sort_values('Peso_Actual', ascending=False)

        sector_group['Target_%'] = sector_group['Sector'].map(st.session_state.sector_targets).fillna(0)
        sector_group['Desvío_%'] = sector_group['Peso_Actual'] - sector_group['Target_%']
        sector_group['Monto_USD'] = sector_group['Monto_USD'].apply(fmt_usd)
        sector_group['Peso_Actual'] = sector_group['Peso_Actual'].round(2)

        st.dataframe(
            sector_group.rename(columns={
                'Sector':'Sector','Monto_USD':'Monto USD','Peso_Actual':'Peso Actual %',
                'Target_%':'Target %','Desvío_%':'Desvío %','Tickers':'Activos'
            }),
            use_container_width=True, hide_index=True
        )

        # Alerts
        for _, row in sector_group.iterrows():
            if row['Target_%'] > 0 and abs(row['Desvío_%']) > 2:
                icon = "📈" if row['Desvío_%'] > 0 else "📉"
                color = "orange" if row['Desvío_%'] > 0 else "blue"
                st.warning(f"{icon} **{row['Sector']}**: sobrecomprado en {row['Desvío_%']:.2f}%" if row['Desvío_%']>0
                    else f"{icon} **{row['Sector']}**: subcomprado en {abs(row['Desvío_%']):.2f}%")

        # Charts
        col1, col2 = st.columns(2)
        with col1:
            sg = df.groupby('Sector')['Peso_Actual_%'].sum().reset_index().sort_values('Peso_Actual_%', ascending=False)
            fig = px.pie(sg, values='Peso_Actual_%', names='Sector',
                title='Composición por Sector (Actual)',
                color_discrete_sequence=COLORS)
            fig.update_traces(textposition='inside', textinfo='label+percent',
                textfont_size=11, hole=0.3)
            fig.update_layout(paper_bgcolor='#111827', plot_bgcolor='#111827',
                font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, showlegend=True,
                legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            sgt = sector_group.copy()
            fig2 = go.Figure()
            fig2.add_trace(go.Bar(name='Actual %', x=sgt['Sector'], y=sgt['Peso_Actual'],
                marker_color='rgba(0,212,255,0.7)', text=sgt['Peso_Actual'].round(1).astype(str)+'%',
                textposition='outside'))
            fig2.add_trace(go.Bar(name='Target %', x=sgt['Sector'], y=sgt['Target_%'],
                marker_color='rgba(255,107,53,0.6)'))
            fig2.update_layout(barmode='group', title='Actual vs Target por Sector',
                paper_bgcolor='#111827', plot_bgcolor='#111827',
                font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='%',
                legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
            st.plotly_chart(fig2, use_container_width=True)

        # ── Fila 2: Desvío por sector + Pareto ──
        st.markdown("---")
        col3, col4 = st.columns(2)

        with col3:
            sg_dev = sector_group.copy()
            sg_dev['Peso_Actual_num'] = df.groupby('Sector')['Peso_Actual_%'].sum().reindex(sg_dev['Sector']).values
            has_targets = (sg_dev['Target_%'] > 0).any()
            if has_targets:
                sg_dev = sg_dev[sg_dev['Target_%'] > 0].copy()
                sg_dev['Desvío_num'] = sg_dev['Peso_Actual_num'] - sg_dev['Target_%']
            else:
                avg = sg_dev['Peso_Actual_num'].mean()
                sg_dev['Desvío_num'] = sg_dev['Peso_Actual_num'] - avg
            # ✅ Fix 3: ordenado de mayor a menor (mismo criterio que Pareto)
            sg_dev = sg_dev.sort_values('Desvío_num', ascending=False)
            colors_dev = ['rgba(74,222,128,0.8)' if v >= 0 else 'rgba(248,113,113,0.8)'
                for v in sg_dev['Desvío_num']]
            fig3 = go.Figure(go.Bar(
                x=sg_dev['Sector'], y=sg_dev['Desvío_num'].round(2),
                marker_color=colors_dev,
                text=sg_dev['Desvío_num'].round(2).astype(str)+'%',
                textposition='outside'
            ))
            fig3.add_hline(y=0, line_dash='dot', line_color='rgba(255,255,255,0.2)')
            fig3.update_layout(
                title='Desvío por Sector (Actual − Target)',
                paper_bgcolor='#111827', plot_bgcolor='#111827',
                font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='Desvío %',
                showlegend=False
            )
            st.plotly_chart(fig3, use_container_width=True)

        with col4:
            sg_par = df.groupby('Sector')['Peso_Actual_%'].sum().reset_index()
            sg_par.columns = ['Sector', 'Peso']
            sg_par = sg_par.sort_values('Peso', ascending=False).reset_index(drop=True)
            sg_par['Acumulado'] = sg_par['Peso'].cumsum().round(2)

            fig4 = go.Figure()
            fig4.add_trace(go.Bar(
                x=sg_par['Sector'], y=sg_par['Peso'].round(2),
                name='Peso %',
                marker_color='rgba(0,212,255,0.7)',
                text=sg_par['Peso'].round(1).astype(str)+'%',
                textposition='outside',
                yaxis='y'
            ))
            fig4.add_trace(go.Scatter(
                x=sg_par['Sector'], y=sg_par['Acumulado'],
                name='Acumulado %',
                mode='lines+markers',
                line=dict(color='#ffffff', width=2),
                marker=dict(size=6, color='white'),
                yaxis='y2'
            ))
            fig4.update_layout(
                title='Concentración por Sector (Pareto)',
                paper_bgcolor='#111827', plot_bgcolor='#111827',
                font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14,
                yaxis=dict(title='Peso %', gridcolor='rgba(30,45,69,0.6)'),
                yaxis2=dict(title='Acum %', overlaying='y', side='right',
                    range=[0, 105], showgrid=False, ticksuffix='%'),
                legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)')
            )
            st.plotly_chart(fig4, use_container_width=True)


# ══════════════════════════════════════════
#  TAB 3 — REBALANCEO
# ══════════════════════════════════════════
with tabs[5]:
    st.subheader("⚖️ Rebalanceo")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    else:
        df = calc_portfolio_weights(st.session_state.portfolio.copy())
        total = df['Total'].iloc[0]

        m1,m2,m3,m4,m5 = st.columns(5)
        m1.metric("Total", fmt_usd(total))
        m2.metric("Activos", len(df))
        m3.metric("A comprar", len(df[df['Desvio_%']<-0.5]), delta="subweight")
        m4.metric("A vender", len(df[df['Desvio_%']>0.5]), delta="overweight")
        vol_mover = df[df['Rebalanceo_USD']>0]['Rebalanceo_USD'].sum()
        m5.metric("Volumen a mover", fmt_usd(vol_mover))

        st.markdown("#### Tabla de rebalanceo (ordenada por acción)")
        df_reb = df.sort_values('Rebalanceo_USD').copy()
        df_reb['Acción'] = df_reb['Rebalanceo_USD'].apply(
            lambda x: '▲ COMPRAR' if x>50 else ('▼ VENDER' if x<-50 else '✓ OK'))

        def color_rebalanceo(val):
            if isinstance(val, (int,float)):
                if val > 50: return 'color: #4ade80; font-weight: bold'
                if val < -50: return 'color: #f87171; font-weight: bold'
            return ''

        display_cols = ['Ticker','Sector','Monto_USD','Target_USD','Rebalanceo_USD',
                        'Peso_Actual_%','Target_%','Desvio_%','Acción']
        st.dataframe(
            df_reb[display_cols].round(2).style.map(color_rebalanceo, subset=['Rebalanceo_USD']),
            use_container_width=True, hide_index=True
        )

        col1, col2 = st.columns(2)
        with col1:
            df_sorted = df.sort_values('Peso_Actual_%', ascending=False)
            fig = go.Figure()
            fig.add_trace(go.Bar(name='Actual %', x=df_sorted['Ticker'], y=df_sorted['Peso_Actual_%'],
                marker_color='rgba(0,212,255,0.75)', text=df_sorted['Peso_Actual_%'].round(1).astype(str)+'%',
                textposition='outside'))
            fig.add_trace(go.Bar(name='Target %', x=df_sorted['Ticker'], y=df_sorted['Target_%'],
                marker_color='rgba(255,107,53,0.65)'))
            fig.update_layout(barmode='group', title='Tenencia Actual vs Target',
                paper_bgcolor='#111827', plot_bgcolor='#111827',
                font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='%', legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            fig2 = px.pie(df, values='Peso_Actual_%', names='Ticker',
                title='Composición actual',
                color_discrete_sequence=px.colors.qualitative.Set3)
            fig2.update_traces(textposition='inside', textinfo='label+percent',
                textfont_size=10, hole=0.25)
            fig2.update_layout(paper_bgcolor='#111827', plot_bgcolor='#111827',
                font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, showlegend=False)
            st.plotly_chart(fig2, use_container_width=True)

        buf = io.BytesIO()
        df_reb[display_cols].round(2).to_excel(buf, index=False)
        st.download_button("⬇️ Exportar rebalanceo Excel", buf.getvalue(), "rebalanceo.xlsx")


# ══════════════════════════════════════════
#  TAB 4 — ESTRATEGIA
# ══════════════════════════════════════════
with tabs[6]:
    st.subheader("🎯 Estrategia de Cartera")
    st.markdown("Definí tu **política de inversión**: cuánto % querés en cada sector, y cómo distribuirlo entre los instrumentos.")

    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero en la pestaña 📁 Mi Cartera")
    else:
        df_port = calc_portfolio_weights(st.session_state.portfolio.copy())
        total   = df_port['Total'].iloc[0]

        # ── Inicializar sector_targets si no existe ──
        if 'sector_targets' not in st.session_state:
            st.session_state.sector_targets = {}

        # Sectores presentes en la cartera
        sectores_cartera = sorted(df_port['Sector'].fillna('Sin sector').unique().tolist())

        # Inicializar targets faltantes desde suma de instrumentos actuales
        for sec in sectores_cartera:
            if sec not in st.session_state.sector_targets:
                pesos_sec = df_port[df_port['Sector']==sec]['Target_%'].sum()
                st.session_state.sector_targets[sec] = round(float(pesos_sec), 1)

        # ─────────────────────────────────────────
        # PASO 1 — Targets por sector
        # ─────────────────────────────────────────
        st.markdown("### 1️⃣ Target por sector")
        st.markdown("Editá el **% objetivo** de cada sector. Al cambiar un sector, los instrumentos se redistribuyen proporcionalmente.")

        total_sector_target = sum(st.session_state.sector_targets.get(s, 0) for s in sectores_cartera)
        if abs(total_sector_target - 100) > 0.1:
            st.warning(f"⚠️ Los targets de sector suman **{total_sector_target:.1f}%** — deben sumar 100%")
        else:
            st.success(f"✅ Targets de sector suman 100%")

        # Sector target editor — una fila por sector
        sec_rows = []
        for sec in sectores_cartera:
            tickers_sec = df_port[df_port['Sector']==sec]['Ticker'].tolist()
            peso_real   = df_port[df_port['Sector']==sec]['Peso_Actual_%'].sum()
            target_sec  = st.session_state.sector_targets.get(sec, 0.0)
            sec_rows.append({
                'Sector': sec,
                'Target Sector %': round(target_sec, 1),
                'Peso Real %': round(peso_real, 2),
                'Diferencia %': round(target_sec - peso_real, 2),
                'Instrumentos': ', '.join(tickers_sec),
            })

        df_sec = pd.DataFrame(sec_rows)

        def color_diff(val):
            try:
                v = float(val)
                if v > 1:  return 'background-color: rgba(74,222,128,0.2); color: #4ade80'
                if v < -1: return 'background-color: rgba(248,113,113,0.2); color: #f87171'
                return 'background-color: rgba(251,191,36,0.15); color: #fbbf24'
            except: return ''

        st.dataframe(
            df_sec.style.map(color_diff, subset=['Diferencia %']),
            use_container_width=True, hide_index=True
        )

        # Sector sliders
        st.markdown("#### Ajustá los targets de sector")
        cols_sec = st.columns(min(4, len(sectores_cartera)))
        changed_sectors = {}
        for i, sec in enumerate(sectores_cartera):
            with cols_sec[i % len(cols_sec)]:
                new_val = st.number_input(
                    f"{sec}", min_value=0.0, max_value=100.0, step=0.5,
                    value=float(st.session_state.sector_targets.get(sec, 0.0)),
                    key=f"sec_target_{sec}"
                )
                changed_sectors[sec] = new_val

        if st.button("💾 Guardar targets de sector y redistribuir", type="primary"):
            # Save sector targets
            st.session_state.sector_targets = changed_sectors

            # Redistribute instrument targets proportionally within each sector
            new_portfolio = st.session_state.portfolio.copy()
            for sec, new_target_sec in changed_sectors.items():
                mask = new_portfolio['Sector'] == sec
                tickers_in_sec = new_portfolio[mask]['Ticker'].tolist()
                if not tickers_in_sec:
                    continue
                # Current instrument targets within sector
                old_targets = new_portfolio[mask]['Target_%'].values.astype(float)
                old_sum = old_targets.sum()
                if old_sum > 0:
                    # Proportional redistribution
                    new_targets = old_targets / old_sum * new_target_sec
                else:
                    # Equal distribution if no existing targets
                    new_targets = np.full(len(old_targets), new_target_sec / len(old_targets))
                new_portfolio.loc[mask, 'Target_%'] = np.round(new_targets, 2)

            st.session_state.portfolio = new_portfolio
            save_persistent(
                st.session_state.portfolio,
                st.session_state.instruments,
                st.session_state.sectors,
                st.session_state.sector_targets
            )
            st.success("✅ Targets guardados y redistribuidos proporcionalmente")
            st.rerun()

        st.markdown("---")

        # ─────────────────────────────────────────
        # PASO 2 — Tabla maestra sector + instrumento
        # ─────────────────────────────────────────
        st.markdown("### 2️⃣ Tabla maestra — Sector → Instrumento")
        st.markdown("Podés ajustar el **target de cada instrumento** directamente. La app te avisa si la suma del sector no coincide con el target del sector.")

        # Build master table
        master_rows = []
        for sec in sectores_cartera:
            df_sec_inst = df_port[df_port['Sector']==sec].copy()
            target_sec  = st.session_state.sector_targets.get(sec, 0.0)
            sum_inst    = df_sec_inst['Target_%'].sum()
            ok          = abs(sum_inst - target_sec) < 0.2

            for _, row in df_sec_inst.iterrows():
                master_rows.append({
                    'Sector': sec,
                    'Target Sector %': round(target_sec, 1),
                    'Ticker': row['Ticker'],
                    'Target Instrumento %': round(row['Target_%'], 2),
                    'Peso Real %': round(row['Peso_Actual_%'], 2),
                    'Desvío %': round(row['Target_%'] - row['Peso_Actual_%'], 2),
                    '✓': '✅' if ok else '⚠️',
                })
            # Subtotal row
            master_rows.append({
                'Sector': f'— Subtotal {sec}',
                'Target Sector %': round(target_sec, 1),
                'Ticker': '',
                'Target Instrumento %': round(sum_inst, 2),
                'Peso Real %': round(df_sec_inst['Peso_Actual_%'].sum(), 2),
                'Desvío %': round(target_sec - df_sec_inst['Peso_Actual_%'].sum(), 2),
                '✓': '✅' if ok else f'⚠️ faltan {target_sec-sum_inst:+.1f}%',
            })

        df_master = pd.DataFrame(master_rows)

        # Editable — only Target Instrumento % is editable
        edited_master = st.data_editor(
            df_master,
            column_config={
                'Sector':                  st.column_config.TextColumn('Sector', disabled=True),
                'Target Sector %':         st.column_config.NumberColumn('Target Sector %', format="%.1f%%", disabled=True),
                'Ticker':                  st.column_config.TextColumn('Ticker', disabled=True),
                'Target Instrumento %':    st.column_config.NumberColumn('Target Instrumento %', format="%.2f%%"),
                'Peso Real %':             st.column_config.NumberColumn('Peso Real %', format="%.2f%%", disabled=True),
                'Desvío %':                st.column_config.NumberColumn('Desvío %', format="%.2f%%", disabled=True),
                '✓':                       st.column_config.TextColumn('Estado', disabled=True),
            },
            hide_index=True,
            use_container_width=True,
            key="master_strategy_editor"
        )

        st.warning("⚠️ **Acordate de guardar** los cambios de targets de instrumento con el botón de abajo.")

        if st.button("💾 Guardar targets de instrumento"):
            # Apply edited instrument targets back to portfolio (skip subtotal rows)
            new_port = st.session_state.portfolio.copy()
            for _, row in edited_master.iterrows():
                ticker = str(row.get('Ticker','')).strip()
                if not ticker or ticker == '':
                    continue  # skip subtotal rows
                if ticker in new_port['Ticker'].values:
                    new_port.loc[new_port['Ticker']==ticker, 'Target_%'] = float(row['Target Instrumento %'])

            # Validate sector sums vs sector targets
            warnings_list = []
            for sec in sectores_cartera:
                target_sec = st.session_state.sector_targets.get(sec, 0.0)
                sum_inst   = new_port[new_port['Sector']==sec]['Target_%'].sum()
                if abs(sum_inst - target_sec) > 0.5:
                    warnings_list.append(f"**{sec}**: instrumentos suman {sum_inst:.1f}% pero target sector es {target_sec:.1f}%")

            if warnings_list:
                for w in warnings_list:
                    st.warning(f"⚠️ {w}")

            st.session_state.portfolio = new_port
            save_persistent(
                st.session_state.portfolio,
                st.session_state.instruments,
                st.session_state.sectors,
                st.session_state.sector_targets
            )
            st.success("✅ Targets de instrumento guardados")
            st.rerun()

        with st.expander("📖 ¿Cómo usar esta pestaña?"):
            st.markdown("""
- **Paso 1**: Definí el % objetivo de cada sector. Al guardar, los instrumentos se redistribuyen proporcionalmente.
- **Paso 2**: Ajustá fino el target de cada instrumento dentro del sector. La columna ✓ te indica si la suma de instrumentos coincide con el target del sector.
- **Verde ✅**: la suma de instrumentos = target del sector. Todo alineado.
- **Amarillo ⚠️**: hay diferencia — los instrumentos no suman el target del sector.
- Los cambios se guardan en disco y persisten entre sesiones.
            """)


# ══════════════════════════════════════════
#  TAB 4 — CORRELACIÓN
# ══════════════════════════════════════════
with tabs[7]:
    st.subheader("🔗 Correlación (datos reales)")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    elif st.session_state.hist_data is None:
        st.warning("⬅️ Hacé click en **Cargar datos de mercado** en el panel lateral")
    else:
        prices = st.session_state.hist_data
        available = [t for t in st.session_state.portfolio['Ticker'].tolist() if t in prices.columns]
        if len(available) < 2:
            st.warning("Se necesitan al menos 2 tickers con datos disponibles")
        else:
            prices_clean = prices[available].ffill().bfill().dropna(how='all')
            returns = prices_clean.pct_change().dropna()
            corr = returns.corr().round(3)

            # Alert high correlations
            high = []
            for i in range(len(corr.columns)):
                for j in range(i+1, len(corr.columns)):
                    val = corr.iloc[i,j]
                    if val > 0.80:
                        high.append(f"{corr.columns[i]}–{corr.columns[j]}: **{val:.2f}**")
            if high:
                st.error(f"⚠️ Alta correlación (>0.80): {' | '.join(high)} — considerá reducir exposición")
            else:
                st.success("✅ Diversificación correcta — ninguna correlación supera 0.80")

            # Heatmap
            fig = px.imshow(corr, text_auto=True, aspect='auto',
                color_continuous_scale='RdYlGn_r',
                title=f'Matriz de Correlación ({years} años de historia)',
                zmin=-1, zmax=1)
            fig.update_layout(paper_bgcolor='#111827', plot_bgcolor='#111827',
                font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, height=500)
            fig.update_traces(textfont_size=11)
            st.plotly_chart(fig, use_container_width=True)

            col1, col2 = st.columns(2)
            with col1:
                avg_corr = corr.apply(lambda x: x[x.index!=x.name].mean()).sort_values(ascending=False)
                colors_corr = ['rgba(248,113,113,.8)' if v>0.6 else 'rgba(251,191,36,.8)' if v>0.4
                    else 'rgba(74,222,128,.8)' for v in avg_corr]
                fig2 = go.Figure(go.Bar(x=avg_corr.index, y=avg_corr.values,
                    marker_color=colors_corr,
                    text=avg_corr.round(2).values, textposition='outside'))
                fig2.update_layout(title='Correlación promedio por activo',
                    paper_bgcolor='#111827', plot_bgcolor='#111827',
                    font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis=dict(range=[0,1]))
                st.plotly_chart(fig2, use_container_width=True)

            with col2:
                # Rolling correlation for top pair
                pairs = [(corr.columns[i], corr.columns[j], corr.iloc[i,j])
                    for i in range(len(corr.columns)) for j in range(i+1,len(corr.columns))]
                pairs.sort(key=lambda x: abs(x[2]), reverse=True)
                if pairs:
                    t1, t2, _ = pairs[0]
                    rolling = returns[t1].rolling(60).corr(returns[t2]).dropna()
                    fig3 = go.Figure(go.Scatter(x=rolling.index, y=rolling.values,
                        fill='tozeroy', line=dict(color='#00d4ff', width=2)))
                    fig3.add_hline(y=0.8, line_dash='dash', line_color='#f87171',
                        annotation_text='Alto (0.80)')
                    fig3.update_layout(title=f'Correlación móvil 60d: {t1} vs {t2}',
                        paper_bgcolor='#111827', plot_bgcolor='#111827',
                        font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='Correlación')
                    st.plotly_chart(fig3, use_container_width=True)

        with st.expander("📖 ¿Cómo interpretar la correlación?"):
            st.markdown("""
- La correlación mide cuánto se mueven juntos dos activos. Va de **-1 a +1**.
- **+1**: se mueven exactamente igual (no diversifican).
- **0**: movimientos independientes (buena diversificación).
- **-1**: se mueven opuesto (cobertura perfecta).
- ⚠️ Correlación **>0.80** es una alerta: esos activos se comportan casi igual y no agregan diversificación real.
- El gráfico de correlación móvil muestra si la relación entre dos activos cambia con el tiempo — lo ideal es que sea estable y baja.
            """)
with tabs[8]:
    st.subheader("📐 Métricas reales")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    elif st.session_state.hist_data is None:
        st.warning("⬅️ Hacé click en **Cargar datos de mercado** en el panel lateral")
    else:
        prices = st.session_state.hist_data
        df_port = calc_portfolio_weights(st.session_state.portfolio.copy())
        available = [t for t in df_port['Ticker'].tolist() if t in prices.columns]
        prices_clean = prices[available].ffill().bfill().dropna(how='all')
        # Drop tickers with fewer than 30 data points
        valid_tickers = [t for t in available if prices_clean[t].notna().sum() >= 30]
        dropped = [t for t in available if t not in valid_tickers]
        if dropped:
            st.warning(f"⚠️ Sin datos suficientes (excluidos del cálculo): {', '.join(dropped)}")
        available = valid_tickers
        if not available:
            st.error("❌ Ningún ticker tiene datos suficientes. Probá aumentar los años de historia en el panel lateral.")
            st.stop()
        prices_clean = prices_clean[available]

        with st.spinner("Calculando métricas..."):
            try:
                metrics, returns = calc_metrics(prices_clean, rf=rf_rate)
            except Exception as e:
                st.error(f"❌ No se pudieron calcular las métricas: {e}\n\nAsegurate de que los tickers tengan datos suficientes y de haber hecho click en **Cargar datos de mercado**.")
                st.stop()

        # Portfolio-level
        weights_actual = np.array([
            df_port[df_port['Ticker']==t]['Peso_Actual_%'].values[0]/100
            for t in available if t in df_port['Ticker'].values
        ])
        if len(weights_actual) == len(available):
            port_ret, port_vol, port_sharpe = portfolio_metrics(weights_actual, returns, rf_rate)
            port_var1d = (-np.dot(weights_actual, returns.quantile(0.05))) * 100
            beta_values = metrics['Beta SPY'].dropna()
            port_beta = float(np.dot(
                [w for t,w in zip(available, weights_actual) if t in metrics.index and not np.isnan(metrics.loc[t,'Beta SPY'])],
                beta_values.reindex([t for t in available if t in beta_values.index]).dropna().values
            )) if len(beta_values) > 0 else np.nan

            m1,m2,m3,m4,m5,m6 = st.columns(6)
            m1.metric("Retorno Anual", f"{port_ret*100:.1f}%")
            m2.metric("Volatilidad", f"{port_vol*100:.1f}%")
            m3.metric("Sharpe", f"{port_sharpe:.2f}")
            m4.metric("Beta vs SPY", f"{port_beta:.2f}" if not np.isnan(port_beta) else "—")
            m5.metric("VaR 1d 95%", f"{port_var1d:.2f}%")
            m6.metric("VaR 10d 95%", f"{port_var1d*np.sqrt(10):.2f}%")
            st.session_state['port_stats'] = {'ret':port_ret,'vol':port_vol,'sharpe':port_sharpe,'beta':port_beta if not np.isnan(port_beta) else 1.0}

        # Table
        st.markdown("#### Métricas por activo")
        metrics_display = metrics.copy()
        metrics_display.index.name = 'Ticker'
        st.dataframe(metrics_display, use_container_width=True)

        col1, col2 = st.columns(2)
        with col1:
            sharpe_sorted = metrics['Sharpe'].sort_values(ascending=True)
            colors_s = ['rgba(74,222,128,.8)' if v>1 else 'rgba(0,212,255,.7)' if v>0.5
                else 'rgba(248,113,113,.8)' for v in sharpe_sorted]
            fig = go.Figure(go.Bar(y=sharpe_sorted.index, x=sharpe_sorted.values,
                orientation='h', marker_color=colors_s,
                text=sharpe_sorted.round(2).values, textposition='outside'))
            fig.update_layout(title='Sharpe Ratio por activo',
                paper_bgcolor='#111827', plot_bgcolor='#111827', font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14)
            st.plotly_chart(fig, use_container_width=True)

        with col2:
            # Risk-Return scatter
            fig2 = go.Figure()
            for i, ticker in enumerate(metrics.index):
                fig2.add_trace(go.Scatter(
                    x=[metrics.loc[ticker,'Volatilidad %']],
                    y=[metrics.loc[ticker,'Retorno Anual %']],
                    mode='markers+text', name=ticker,
                    text=[ticker], textposition='top center',
                    marker=dict(size=12, color=COLORS[i % len(COLORS)])
                ))
            fig2.update_layout(title='Mapa Riesgo vs Retorno',
                paper_bgcolor='#111827', plot_bgcolor='#111827',
                font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, showlegend=False,
                xaxis_title='Volatilidad %', yaxis_title='Retorno Anual %')
            st.plotly_chart(fig2, use_container_width=True)

        with st.expander("📖 ¿Cómo interpretar estas métricas?"):
            st.markdown("""
- **Retorno Anual %**: Rendimiento anualizado promedio en el período seleccionado.
- **Volatilidad %**: Desviación estándar anualizada — mide el riesgo total del activo.
- **Sharpe**: Retorno ajustado por riesgo. >1 es bueno, >2 es excelente. Menor a 0.5 indica que el riesgo no se está compensando.
- **CAGR %**: Tasa de crecimiento anual compuesta — el retorno real si hubieras mantenido el activo todo el período.
- **Beta SPY**: Sensibilidad al mercado. Beta=1 se mueve igual que el mercado. Beta>1 amplifica movimientos. Beta<1 es más defensivo.
- **VaR 1d 95%**: Pérdida máxima esperada en un día normal (95% de confianza). Ejemplo: VaR=2% significa que solo hay 5% de chances de perder más de 2% en un día.
            """)


# ══════════════════════════════════════════
#  TAB 6 — RENDIMIENTO
# ══════════════════════════════════════════
with tabs[9]:
    st.subheader("📈 Rendimiento acumulado")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    elif st.session_state.hist_data is None:
        st.warning("⬅️ Cargá datos de mercado primero")
    else:
        prices = st.session_state.hist_data
        df_port = calc_portfolio_weights(st.session_state.portfolio.copy())
        available = [t for t in df_port['Ticker'].tolist() if t in prices.columns]
        prices_clean = prices[available].ffill().bfill().dropna(how='all')

        weights_actual = np.array([df_port[df_port['Ticker']==t]['Peso_Actual_%'].values[0]/100
            for t in available])

        # Portfolio combined price
        port_prices = (prices_clean * weights_actual).sum(axis=1)
        port_cum = (port_prices / port_prices.iloc[0] - 1) * 100

        # Download benchmarks
        with st.spinner("Descargando SPY y QQQ..."):
            try:
                bench_raw = yf.download(['SPY','QQQ'], period=PERIOD, auto_adjust=True, progress=False)
                if isinstance(bench_raw.columns, pd.MultiIndex):
                    bench = bench_raw['Close']
                else:
                    bench = bench_raw
                bench = bench.reindex(port_cum.index, method='ffill').dropna()
                spy_cum  = (bench['SPY']  / bench['SPY'].iloc[0]  - 1) * 100 if 'SPY'  in bench.columns else None
                qqq_cum  = (bench['QQQ']  / bench['QQQ'].iloc[0]  - 1) * 100 if 'QQQ'  in bench.columns else None
            except:
                spy_cum, qqq_cum = None, None

        fig = go.Figure()
        fig.add_trace(go.Scatter(x=port_cum.index, y=port_cum.values,
            name='Mi Cartera', line=dict(color='#00d4ff', width=2.5)))
        if spy_cum is not None:
            fig.add_trace(go.Scatter(x=spy_cum.index, y=spy_cum.values,
                name='SPY', line=dict(color='#f87171', width=1.5, dash='dash')))
        if qqq_cum is not None:
            fig.add_trace(go.Scatter(x=qqq_cum.index, y=qqq_cum.values,
                name='QQQ', line=dict(color='#fbbf24', width=1.5, dash='dot')))

        # Individual tickers
        show_tickers = st.checkbox("Mostrar activos individuales", value=False)
        if show_tickers:
            for i, t in enumerate(available):
                cum = (prices_clean[t] / prices_clean[t].iloc[0] - 1) * 100
                fig.add_trace(go.Scatter(x=cum.index, y=cum.values,
                    name=t, line=dict(width=1, dash='dot'),
                    opacity=0.6))

        fig.add_hline(y=0, line_dash='dot', line_color='rgba(255,255,255,0.2)')
        fig.update_layout(
            title='Rendimiento Acumulado (%)',
            paper_bgcolor='#111827', plot_bgcolor='#111827',
            font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='Retorno %',
            legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)')
        )
        st.plotly_chart(fig, use_container_width=True)

        # CAGR comparison bar
        st.markdown("#### CAGR comparativo")
        cagr_items = {'Mi Cartera': float((port_prices.iloc[-1]/port_prices.iloc[0])**(252/len(port_prices))-1)*100}
        if spy_cum is not None:
            cagr_items['SPY'] = float((bench['SPY'].iloc[-1]/bench['SPY'].iloc[0])**(252/len(bench))-1)*100
        if qqq_cum is not None:
            cagr_items['QQQ'] = float((bench['QQQ'].iloc[-1]/bench['QQQ'].iloc[0])**(252/len(bench))-1)*100

        fig2 = go.Figure(go.Bar(
            x=list(cagr_items.keys()), y=list(cagr_items.values()),
            marker_color=['#00d4ff','#f87171','#fbbf24'][:len(cagr_items)],
            text=[f"{v:.1f}%" for v in cagr_items.values()], textposition='outside'
        ))
        fig2.update_layout(title='CAGR anualizado',
            paper_bgcolor='#111827', plot_bgcolor='#111827',
            font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='%', legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
        st.plotly_chart(fig2, use_container_width=True)

        with st.expander("📖 ¿Cómo interpretar este gráfico?"):
            st.markdown("""
- El gráfico muestra el crecimiento acumulado de $1 invertido al inicio del período.
- **Línea azul** = tu cartera con los pesos actuales.
- **Línea roja** = SPY (S&P 500 — benchmark del mercado americano).
- **Línea amarilla** = QQQ (Nasdaq 100 — benchmark tecnológico).
- **CAGR** = Tasa de Crecimiento Anual Compuesta — es el retorno anual equivalente si hubieras mantenido la inversión todo el período.
            """)


# ══════════════════════════════════════════
#  TAB 7 — VaR
# ══════════════════════════════════════════
with tabs[10]:
    st.subheader("⚠️ Value at Risk (VaR) — 95% de confianza")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    elif st.session_state.hist_data is None:
        st.warning("⬅️ Cargá datos de mercado primero")
    else:
        prices = st.session_state.hist_data
        df_port = calc_portfolio_weights(st.session_state.portfolio.copy())
        available = [t for t in df_port['Ticker'].tolist() if t in prices.columns]
        prices_clean = prices[available].ffill().bfill().dropna(how='all')
        returns_all = prices_clean.pct_change().dropna()

        weights_actual = np.array([df_port[df_port['Ticker']==t]['Peso_Actual_%'].values[0]/100
            for t in available])

        port_ret_series = returns_all.dot(weights_actual)
        var_1d = float(-port_ret_series.quantile(0.05) * 100)
        var_10d = var_1d * np.sqrt(10)
        cvar = float(-port_ret_series[port_ret_series <= port_ret_series.quantile(0.05)].mean() * 100)

        total = df_port['Total'].iloc[0]
        m1,m2,m3,m4 = st.columns(4)
        m1.metric("VaR 1d 95%", f"{var_1d:.2f}%", delta=f"-{var_1d/100*total:,.0f} USD")
        m2.metric("VaR 10d 95%", f"{var_10d:.2f}%", delta=f"-{var_10d/100*total:,.0f} USD")
        m3.metric("CVaR 1d 95%", f"{cvar:.2f}%", delta=f"-{cvar/100*total:,.0f} USD")
        m4.metric("Capital en riesgo (1d)", fmt_usd(var_1d/100*total))

        # Histogram
        fig = go.Figure()
        fig.add_trace(go.Histogram(
            x=port_ret_series.values * 100,
            nbinsx=60,
            name='Retornos diarios',
            marker_color='rgba(0,212,255,0.6)',
            opacity=0.8
        ))
        fig.add_vline(x=-var_1d, line_dash='dash', line_color='#f87171', line_width=2,
            annotation_text=f'VaR 95% = -{var_1d:.2f}%',
            annotation_font_color='#f87171', annotation_position='top left')
        fig.add_vline(x=-cvar, line_dash='dot', line_color='#fbbf24', line_width=2,
            annotation_text=f'CVaR = -{cvar:.2f}%',
            annotation_font_color='#fbbf24', annotation_position='bottom left')
        fig.update_layout(
            title='Distribución de Retornos Diarios de la Cartera',
            paper_bgcolor='#111827', plot_bgcolor='#111827',
            font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, xaxis_title='Retorno Diario %', yaxis_title='Frecuencia',
            showlegend=False
        )
        st.plotly_chart(fig, use_container_width=True)

        # VaR by ticker
        st.markdown("#### VaR por activo")
        var_data = []
        for t in available:
            r = returns_all[t]
            v1 = float(-r.quantile(0.05)*100)
            var_data.append({'Ticker':t, 'VaR 1d %':round(v1,2), 'VaR 10d %':round(v1*np.sqrt(10),2),
                'CVaR 1d %':round(float(-r[r<=r.quantile(0.05)].mean()*100),2)})
        st.dataframe(pd.DataFrame(var_data).sort_values('VaR 1d %', ascending=False),
            use_container_width=True, hide_index=True)

        with st.expander("📖 ¿Cómo interpretar el VaR?"):
            st.markdown("""
- **VaR 1d 95%**: Pérdida máxima esperada en un **día normal**. Con 95% de confianza, no se espera perder más de ese % en un solo día.
- **VaR 10d 95%**: Mismo concepto pero proyectado a 10 días hábiles (se escala por √10).
- **CVaR (Expected Shortfall)**: Promedio de las pérdidas que **sí superan** el VaR — mide qué tan malo puede ser el peor 5% de los días.
- La línea roja en el histograma marca el VaR. Todo lo que queda a la izquierda de esa línea es el 5% de peores días históricos.
- Ejemplo: VaR=2.5% significa que solo 1 de cada 20 días se perdió más de 2.5%.
            """)



# ══════════════════════════════════════════
#  TAB 6 — STRESS TEST
# ══════════════════════════════════════════
with tabs[11]:
    st.subheader("🔥 Stress Test")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    elif st.session_state.hist_data is None:
        st.warning("⬅️ Cargá datos de mercado primero")
    else:
        port_stats = st.session_state.get('port_stats', {'ret':0.18,'vol':0.22,'beta':1.1})
        df_port = calc_portfolio_weights(st.session_state.portfolio.copy())
        total = df_port['Total'].iloc[0]
        beta = port_stats.get('beta', 1.1)

        st.markdown(f"**Beta de tu cartera vs SPY:** {beta:.2f}")
        st.markdown("---")

        scenarios = [
            ("Corrección leve", -0.05, "2022 drawdown"),
            ("Corrección moderada", -0.10, "Corrección típica"),
            ("Crash severo", -0.20, "Bear market"),
            ("Crisis 2008 (Lehman)", -0.38, "Histórico"),
            ("COVID-19 Crash (Feb-Mar 2020)", -0.34, "Histórico"),
            ("Flash Crash hipotético", -0.50, "Extremo"),
        ]

        st.markdown("#### Escenarios hipotéticos")
        data_stress = []
        for label, spy_drop, tipo in scenarios:
            impact = spy_drop * beta
            usd_loss = impact * total
            data_stress.append({
                'Escenario': label,
                'Tipo': tipo,
                'SPY cae': f"{spy_drop*100:.0f}%",
                'Impacto cartera': f"{impact*100:.2f}%",
                'Pérdida USD': fmt_usd(abs(usd_loss)),
                '_impact_val': impact
            })

        df_stress = pd.DataFrame(data_stress)

        def color_stress(val):
            if isinstance(val, str) and '%' in val and val.startswith('-'):
                try:
                    v = float(val.replace('%',''))
                    if v < -30: return 'background-color: rgba(220,38,38,0.4); color: #fca5a5'
                    if v < -15: return 'background-color: rgba(234,88,12,0.3); color: #fdba74'
                    return 'background-color: rgba(202,138,4,0.2); color: #fde68a'
                except: pass
            return ''

        st.dataframe(
            df_stress[['Escenario','Tipo','SPY cae','Impacto cartera','Pérdida USD']]\
                .style.map(color_stress, subset=['Impacto cartera']),
            use_container_width=True, hide_index=True
        )

        # Chart
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=df_stress['Escenario'],
            y=[float(r['_impact_val'])*100 for _,r in df_stress.iterrows()],
            marker_color=[f'rgba(220,38,38,{min(0.9,abs(r["_impact_val"])*2+0.2)})' for _,r in df_stress.iterrows()],
            text=[r['Impacto cartera'] for _,r in df_stress.iterrows()],
            textposition='outside'
        ))
        fig.update_layout(title='Impacto estimado por escenario',
            paper_bgcolor='#111827', plot_bgcolor='#111827',
            font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='% impacto', xaxis_tickangle=-25)
        st.plotly_chart(fig, use_container_width=True)

        # Historical drawdown from real data
        if st.session_state.hist_data is not None:
            st.markdown("#### Drawdown histórico real de tu cartera")
            prices = st.session_state.hist_data
            df_p = calc_portfolio_weights(st.session_state.portfolio.copy())
            available = [t for t in df_p['Ticker'].tolist() if t in prices.columns]
            if available:
                w = np.array([df_p[df_p['Ticker']==t]['Peso_Actual_%'].values[0]/100
                    for t in available])
                port_prices = (prices[available] * w).sum(axis=1)
                port_prices = port_prices / port_prices.iloc[0] * 100
                rolling_max = port_prices.cummax()
                drawdown = (port_prices - rolling_max) / rolling_max * 100

                fig2 = go.Figure(go.Scatter(x=drawdown.index, y=drawdown.values,
                    fill='tozeroy', line=dict(color='#f87171', width=1.5),
                    fillcolor='rgba(248,113,113,0.15)'))
                fig2.update_layout(title='Drawdown histórico real',
                    paper_bgcolor='#111827', plot_bgcolor='#111827',
                    font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='Drawdown %')
                st.plotly_chart(fig2, use_container_width=True)
                st.info(f"📉 Máximo drawdown histórico: **{drawdown.min():.2f}%**")

        with st.expander("📖 ¿Cómo interpretar el Stress Test?"):
            st.markdown("""
- Cada fila muestra la **caída estimada de tu cartera** si el SPY (mercado) baja ese porcentaje.
- El cálculo usa la **Beta** de tu cartera: si Beta=1.2 y el SPY cae 10%, tu cartera caería aproximadamente 12%.
- 🟡 Pérdida leve (<15%) · 🟠 Pérdida moderada (15-30%) · 🔴 Pérdida severa (>30%)
- Los escenarios históricos (2008, COVID) reflejan caídas reales del mercado en esos eventos.
- El **Drawdown histórico** muestra la caída máxima real que tuvo tu cartera en el período analizado.
- ⚠️ Estos son estimados — la relación real puede variar según correlaciones y condiciones de mercado.
            """)


# ══════════════════════════════════════════
#  TAB 9 — FRONTERA EFICIENTE
# ══════════════════════════════════════════
with tabs[12]:
    st.subheader("🌐 Frontera Eficiente de Markowitz")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    elif st.session_state.hist_data is None:
        st.warning("⬅️ Cargá datos de mercado primero")
    else:
        prices = st.session_state.hist_data
        df_port = calc_portfolio_weights(st.session_state.portfolio.copy())
        available = [t for t in df_port['Ticker'].tolist() if t in prices.columns]
        if len(available) < 2:
            st.warning("Se necesitan al menos 2 activos con datos.")
        else:
            prices_clean = prices[available].ffill().bfill().dropna(how='all')
            returns_fe = prices_clean.pct_change().dropna()
            min_w = st.session_state.get('min_weight', 0.0)

            with st.spinner("Calculando frontera eficiente (puede tardar ~20 segundos)..."):
                # Random portfolios
                n = len(available)
                n_sim = 3000
                rand_rets, rand_vols, rand_sharpes, rand_weights = [], [], [], []
                np.random.seed(42)
                for _ in range(n_sim):
                    w = np.random.dirichlet(np.ones(n))
                    r, v, s = portfolio_metrics(w, returns_fe, rf_rate)
                    rand_rets.append(r*100); rand_vols.append(v*100); rand_sharpes.append(s)

                # Key portfolios
                w_ms  = max_sharpe(returns_fe, rf_rate, min_w)
                w_mv  = min_variance(returns_fe, min_w)
                w_act = np.array([df_port[df_port['Ticker']==t]['Peso_Actual_%'].values[0]/100 for t in available])

                ms_r, ms_v, ms_s = portfolio_metrics(w_ms,  returns_fe, rf_rate)
                mv_r, mv_v, mv_s = portfolio_metrics(w_mv,  returns_fe, rf_rate)
                ac_r, ac_v, ac_s = portfolio_metrics(w_act, returns_fe, rf_rate)

            fig = go.Figure()

            # Random cloud
            fig.add_trace(go.Scatter(
                x=rand_vols, y=rand_rets,
                mode='markers',
                marker=dict(size=4, color=rand_sharpes, colorscale='Viridis',
                    showscale=True, colorbar=dict(title='Sharpe', tickfont=dict(color='#e2e8f0')),
                    opacity=0.5),
                name='Portfolios aleatorios', hovertemplate='Vol: %{x:.1f}%<br>Ret: %{y:.1f}%<extra></extra>'
            ))

            # Key points
            for label, rv, rr, rs, color, sym in [
                ('Max Sharpe',    ms_v*100, ms_r*100, ms_s, '#f87171', 'star'),
                ('Min Varianza',  mv_v*100, mv_r*100, mv_s, '#4ade80', 'circle'),
                ('Tu Cartera',    ac_v*100, ac_r*100, ac_s, '#00d4ff', 'x'),
            ]:
                fig.add_trace(go.Scatter(
                    x=[rv], y=[rr], mode='markers+text',
                    name=f'{label} (Sharpe: {rs:.2f})',
                    text=[label], textposition='top center',
                    marker=dict(size=18, color=color, symbol=sym, line=dict(width=2, color=color))
                ))

            fig.update_layout(
                title='Espacio de Portfolios — Frontera Eficiente de Markowitz',
                paper_bgcolor='#111827', plot_bgcolor='#111827', font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14,
                xaxis_title='Volatilidad Anual %', yaxis_title='Retorno Anual %',
                legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'),
                height=550
            )
            st.plotly_chart(fig, use_container_width=True)

            # Composition of key portfolios
            st.markdown("#### Composición de portfolios óptimos")
            col1, col2, col3 = st.columns(3)
            for col, label, w, r, v, s, color in [
                (col1, '🔴 Max Sharpe',   w_ms,  ms_r, ms_v, ms_s, '#f87171'),
                (col2, '🟢 Min Varianza', w_mv,  mv_r, mv_v, mv_s, '#4ade80'),
                (col3, '🔵 Tu Cartera',   w_act, ac_r, ac_v, ac_s, '#00d4ff'),
            ]:
                with col:
                    st.markdown(f"**{label}**")
                    st.metric("Retorno Anual", f"{r*100:.1f}%")
                    st.metric("Volatilidad", f"{v*100:.1f}%")
                    st.metric("Sharpe", f"{s:.2f}")
                    df_comp = pd.DataFrame({'Ticker': available, 'Peso %': (w*100).round(1)})\
                        .sort_values('Peso %', ascending=False)
                    # Convert hex color to rgba for transparency
                    color_map = {'#f87171':'rgba(248,113,113,0.7)','#4ade80':'rgba(74,222,128,0.7)','#00d4ff':'rgba(0,212,255,0.7)'}
                    bar_color = color_map.get(color, 'rgba(0,212,255,0.7)')
                    fig_comp = go.Figure(go.Bar(
                        x=df_comp['Ticker'], y=df_comp['Peso %'],
                        marker_color=bar_color,
                        text=df_comp['Peso %'].astype(str)+'%', textposition='outside'
                    ))
                    fig_comp.update_layout(paper_bgcolor='#111827', plot_bgcolor='#111827',
                        font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, height=250, margin=dict(t=10,b=10),
                        yaxis_title='%', showlegend=False)
                    st.plotly_chart(fig_comp, use_container_width=True)

            with st.expander("📖 ¿Cómo interpretar la Frontera Eficiente?"):
                st.markdown("""
- Cada punto del gráfico representa un **portfolio posible** con distintas combinaciones de pesos.
- El **color** de cada punto indica su Sharpe Ratio (más amarillo = mejor relación riesgo/retorno).
- La **curva superior izquierda** de puntos forma la Frontera Eficiente — son los portfolios que maximizan retorno para cada nivel de riesgo.
- **Max Sharpe** ⭐: el portfolio con la mejor relación retorno/riesgo.
- **Min Varianza** 🟢: el portfolio más conservador — menor volatilidad posible.
- **Tu Cartera** 🔵: dónde está tu distribución actual en este espacio.
- Si tu cartera está lejos de la frontera, existe un portfolio más eficiente con el mismo riesgo pero mayor retorno.
                """)



# ══════════════════════════════════════════
#  TAB 7 — BLACK-LITTERMAN
# ══════════════════════════════════════════
with tabs[13]:
    st.subheader("🧠 Black-Litterman")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    elif st.session_state.hist_data is None:
        st.warning("⬅️ Cargá datos de mercado primero")
    else:
        prices = st.session_state.hist_data
        df_port = calc_portfolio_weights(st.session_state.portfolio.copy())
        available = [t for t in df_port['Ticker'].tolist() if t in prices.columns]
        prices_clean = prices[available].ffill().bfill().dropna(how='all')
        returns = prices_clean.pct_change().dropna()

        st.markdown("#### Ingresá tus views (expectativas de retorno)")
        st.caption("q % anual = tu expectativa de retorno anual para ese activo | Confianza 0–1 = cuánto confiás en ese view")

        views_q, views_conf = {}, {}
        cols = st.columns(3)
        for i, ticker in enumerate(available):
            with cols[i % 3]:
                st.markdown(f"**{ticker}**")
                c1, c2 = st.columns(2)
                views_q[ticker] = c1.number_input(f"q % {ticker}", value=15.0, step=1.0, key=f"q_{ticker}") / 100
                views_conf[ticker] = c2.number_input(f"conf {ticker}", value=0.7, min_value=0.0, max_value=1.0, step=0.05, key=f"c_{ticker}")

        if st.button("🧠 Correr Black-Litterman", type="primary"):
            w_actual = np.array([df_port[df_port['Ticker']==t]['Peso_Actual_%'].values[0]/100
                for t in available])

            posterior_mu = black_litterman(w_actual, returns, views_q, views_conf, rf=rf_rate)
            prior_mu = returns.mean() * 252

            # Read sidebar settings
            min_w = st.session_state.get('min_weight', 0.0)
            target_ret = st.session_state.get('target_return', None)

            # Optimal BL weights
            w_bl = max_sharpe(returns, rf=rf_rate, min_weight=min_w)
            w_minvar = min_variance(returns, min_weight=min_w)

            # Target return portfolio (if enabled)
            w_target = None
            if target_ret is not None:
                w_target = target_return_weights(returns, target_ret, min_weight=min_w)
                if w_target is None:
                    st.warning(f"⚠️ No se encontró solución para retorno objetivo {target_ret*100:.1f}% — puede estar fuera del rango alcanzable con estos activos.")

            # Results table
            bl_df = pd.DataFrame({
                'Ticker': available,
                'Prior (eq.) %': (prior_mu.values * 100).round(2),
                'Posterior BL %': (posterior_mu.values * 100).round(2),
                'Delta %': ((posterior_mu.values - prior_mu.values) * 100).round(2),
                'Peso BL %': (w_bl * 100).round(2),
                'Peso Actual %': (w_actual * 100).round(2),
            }).sort_values('Delta %', ascending=False)

            st.dataframe(bl_df, use_container_width=True, hide_index=True)

            col1, col2 = st.columns(2)
            with col1:
                colors_bl = ['rgba(74,222,128,.8)' if v>=0 else 'rgba(248,113,113,.8)'
                    for v in bl_df['Delta %']]
                fig = go.Figure(go.Bar(x=bl_df['Ticker'], y=bl_df['Delta %'],
                    marker_color=colors_bl, text=bl_df['Delta %'].astype(str)+'%',
                    textposition='outside'))
                fig.update_layout(title='Delta BL (Posterior − Prior)',
                    paper_bgcolor='#111827', plot_bgcolor='#111827',
                    font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='puntos % anual', legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
                st.plotly_chart(fig, use_container_width=True)

            with col2:
                fig2 = go.Figure()
                fig2.add_trace(go.Bar(name='Actual %', x=bl_df['Ticker'],
                    y=bl_df['Peso Actual %'], marker_color='rgba(0,212,255,.7)'))
                fig2.add_trace(go.Bar(name='BL Optimo %', x=bl_df['Ticker'],
                    y=bl_df['Peso BL %'], marker_color='rgba(255,107,53,.7)'))
                fig2.update_layout(barmode='group', title='Pesos: Actual vs BL Óptimo',
                    paper_bgcolor='#111827', plot_bgcolor='#111827',
                    font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14, yaxis_title='%', legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
                st.plotly_chart(fig2, use_container_width=True)

            # Frontier scatter
            port_actual = portfolio_metrics(w_actual, returns, rf_rate)
            port_bl = portfolio_metrics(w_bl, returns, rf_rate)
            port_mv = portfolio_metrics(w_minvar, returns, rf_rate)

            scatter_ports = [
                ('Actual',      port_actual, '#00d4ff', 'x'),
                ('BL Óptimo',   port_bl,     '#f87171', 'square'),
                ('Min Varianza',port_mv,     '#4ade80', 'circle'),
            ]
            if w_target is not None:
                port_tgt = portfolio_metrics(w_target, returns, rf_rate)
                scatter_ports.append((f'Ret.Obj {target_ret*100:.0f}%', port_tgt, '#fbbf24', 'diamond'))

            fig3 = go.Figure()
            for label, stats, color, symbol in scatter_ports:
                fig3.add_trace(go.Scatter(x=[stats[1]*100], y=[stats[0]*100],
                    mode='markers+text', name=f"{label} (Sharpe:{stats[2]:.2f})",
                    text=[label], textposition='top center',
                    marker=dict(size=16, color=color, symbol=symbol,
                        line=dict(width=2, color=color))))
            fig3.update_layout(title='Frontera Eficiente: Actual vs BL vs MinVar',
                paper_bgcolor='#111827', plot_bgcolor='#111827', font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14,
                xaxis_title='Volatilidad %', yaxis_title='Retorno %',
                legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
            st.plotly_chart(fig3, use_container_width=True)

            # Store for MC — prefer target return portfolio if set
            best = port_tgt if (w_target is not None) else port_bl
            st.session_state['bl_stats'] = {
                'ret_bl': best[0], 'vol_bl': best[1]
            }

            with st.expander("📖 ¿Cómo interpretar Black-Litterman?"):
                st.markdown("""
- **Black-Litterman** combina el equilibrio del mercado con tus propias expectativas (views) para calcular pesos óptimos.
- **Prior (equilibrio)**: es el retorno implícito que el mercado "espera" de cada activo basado en su comportamiento histórico.
- **Posterior BL**: es el retorno ajustado incorporando tu view. Si confiás 100% en tu view, el posterior = tu expectativa. Si confiás 0%, el posterior = el prior del mercado.
- **Delta %**: cuánto cambió la expectativa de retorno al incorporar tu view. Verde = subió, rojo = bajó.
- **Peso BL %**: la asignación óptima que maximiza el Sharpe con los retornos posteriores. Cuanto mayor el retorno esperado ajustado y menor la correlación con el resto, mayor será el peso sugerido.
- **Confianza 0–1**: 0 = ignorás tu view completamente, 1 = confiás ciegamente en tu expectativa.
- 💡 Tip: empezá con confianza 0.5–0.7 y ajustá según qué tan seguro estás de tu análisis.
                """)


# ══════════════════════════════════════════
#  TAB 8 — MONTE CARLO
# ══════════════════════════════════════════
with tabs[14]:
    st.subheader("🎲 Monte Carlo")
    if st.session_state.portfolio.empty:
        st.info("Cargá tu cartera primero")
    elif st.session_state.hist_data is None:
        st.warning("⬅️ Cargá datos de mercado primero")
    else:
        df_port = calc_portfolio_weights(st.session_state.portfolio.copy())
        total = df_port['Total'].iloc[0]
        port_stats = st.session_state.get('port_stats', {'ret':0.18,'vol':0.22})
        bl_stats = st.session_state.get('bl_stats', None)

        col1, col2, col3 = st.columns(3)
        n_sims = col1.number_input("N° simulaciones", value=500, min_value=100, max_value=2000, step=100)
        n_days = col2.number_input("Días de trading", value=252, min_value=60, max_value=756, step=21)
        show_paths = col3.number_input("Trayectorias a mostrar", value=60, min_value=10, max_value=200, step=10)

        if st.button("🎲 Correr simulación", type="primary"):
            mu_act = port_stats['ret']
            vol_act = port_stats['vol']
            mu_bl = bl_stats['ret_bl'] if bl_stats else mu_act * 1.15
            vol_bl = bl_stats['vol_bl'] if bl_stats else vol_act * 0.88

            with st.spinner("Simulando..."):
                paths_act = monte_carlo(mu_act, vol_act, total, int(n_sims), int(n_days))
                paths_bl  = monte_carlo(mu_bl,  vol_bl,  total, int(n_sims), int(n_days))

            mean_act = paths_act.mean(axis=0)
            mean_bl  = paths_bl.mean(axis=0)
            days_x = list(range(int(n_days)+1))

            # MC chart
            fig = go.Figure()
            step = max(1, int(n_sims)//int(show_paths))
            for i in range(0, int(n_sims), step):
                fig.add_trace(go.Scatter(x=days_x, y=paths_act[i], mode='lines',
                    line=dict(color='rgba(0,212,255,0.06)', width=1), showlegend=False))
                fig.add_trace(go.Scatter(x=days_x, y=paths_bl[i], mode='lines',
                    line=dict(color='rgba(255,107,53,0.06)', width=1), showlegend=False))
            fig.add_trace(go.Scatter(x=days_x, y=mean_act, mode='lines', name='Media Actual',
                line=dict(color='#00d4ff', width=3)))
            fig.add_trace(go.Scatter(x=days_x, y=mean_bl, mode='lines', name='Media BL Optimizada',
                line=dict(color='#ff6b35', width=3)))
            fig.update_layout(
                title=f'Proyección Monte Carlo ({int(n_days)} días) — Capital inicial: {fmt_usd(total)}',
                paper_bgcolor='#111827', plot_bgcolor='#111827', font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14,
                xaxis_title='Días de Trading', yaxis_title='Valor USD',
                yaxis=dict(tickformat='$,.0f'),
                legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
            st.plotly_chart(fig, use_container_width=True)

            final_act = paths_act[:, -1]
            final_bl  = paths_bl[:, -1]
            med_act = final_act.mean()
            med_bl  = final_bl.mean()
            gain = (med_bl - med_act) / med_act * 100

            m1,m2,m3,m4,m5,m6 = st.columns(6)
            m1.metric("Capital inicial", fmt_usd(total))
            m2.metric("Media Actual (1 año)", fmt_usd(med_act))
            m3.metric("Media BL Optimizada", fmt_usd(med_bl), delta=f"+{gain:.1f}%")
            m4.metric("Peor 5% Actual", fmt_usd(np.percentile(final_act,5)))
            m5.metric("Peor 5% BL", fmt_usd(np.percentile(final_bl,5)))
            m6.metric("Mejor 95% BL", fmt_usd(np.percentile(final_bl,95)))

            col1, col2 = st.columns(2)
            with col1:
                all_vals = np.concatenate([final_act, final_bl])
                bins = np.linspace(all_vals.min(), all_vals.max(), 35)
                hist_act, _ = np.histogram(final_act, bins=bins)
                hist_bl,  _ = np.histogram(final_bl,  bins=bins)
                bin_centers = (bins[:-1] + bins[1:]) / 2

                fig2 = go.Figure()
                fig2.add_trace(go.Bar(x=bin_centers, y=hist_act,
                    name=f'Actual (Media: {fmt_usd(med_act)})',
                    marker_color='rgba(0,212,255,0.55)'))
                fig2.add_trace(go.Bar(x=bin_centers, y=hist_bl,
                    name=f'BL Optim (Media: {fmt_usd(med_bl)})',
                    marker_color='rgba(255,107,53,0.55)'))
                fig2.update_layout(barmode='overlay', title='Distribución de valor final',
                    paper_bgcolor='#111827', plot_bgcolor='#111827', font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14,
                    xaxis=dict(tickformat='$,.0f'), yaxis_title='Frecuencia',
                    legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
                st.plotly_chart(fig2, use_container_width=True)

            with col2:
                pcts = [5,10,25,50,75,90,95]
                fig3 = go.Figure()
                fig3.add_trace(go.Bar(x=[f'P{p}' for p in pcts],
                    y=[np.percentile(final_act,p) for p in pcts],
                    name='Actual', marker_color='rgba(0,212,255,0.7)'))
                fig3.add_trace(go.Bar(x=[f'P{p}' for p in pcts],
                    y=[np.percentile(final_bl,p) for p in pcts],
                    name='BL Optim', marker_color='rgba(255,107,53,0.7)'))
                fig3.update_layout(barmode='group', title='Percentiles de resultado',
                    paper_bgcolor='#111827', plot_bgcolor='#111827', font_color='#e2e8f0', title_font_color='#00d4ff', title_font_size=14,
                    yaxis=dict(tickformat='$,.0f'),
                    legend=dict(font=dict(size=11, color='#e2e8f0'), bgcolor='rgba(0,0,0,0)'))
                st.plotly_chart(fig3, use_container_width=True)

            with st.expander("📖 ¿Cómo interpretar Monte Carlo?"):
                st.markdown("""
- La simulación genera **500 trayectorias posibles** de tu cartera durante 1 año (252 días hábiles), usando el retorno y volatilidad históricos.
- **Línea azul** = trayectoria promedio de tu cartera actual.
- **Línea naranja** = trayectoria promedio de la cartera optimizada con Black-Litterman.
- Las trayectorias tenues de fondo muestran la dispersión real de escenarios posibles — arriba y abajo.
- **Distribución final**: el histograma muestra en cuántas simulaciones terminaste en cada rango de valor. Una distribución más ancha = más incertidumbre.
- **Percentiles**:
  - P5 = peor escenario probable (solo 5% de simulaciones terminaron peor que esto)
  - P50 = resultado mediano esperado
  - P95 = mejor escenario probable (solo 5% terminaron mejor que esto)
- 💡 La diferencia entre la cartera Actual y BL Optimizada refleja el impacto de los pesos sugeridos por Black-Litterman.
- ⚠️ Monte Carlo asume retornos normales y distribución estable — en crisis reales, las colas son más gruesas.
                """)
